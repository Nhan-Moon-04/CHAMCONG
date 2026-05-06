from datetime import date, datetime

from io import BytesIO

from flask import flash, redirect, render_template, request, send_file, session, url_for
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import func

from .database import db
from .models import (
    AttendanceDetail,
    Employee,
    UnionHdRule,
    UnionHolidayEvent,
    UnionHolidayRecipient,
    UnionLedgerEntry,
    UnionYearConfig,
    WorkSchedule,
)
from .services.attendance import parse_month_key
from .services.audit import log_action


DEFAULT_UNION_HD_RULES = [
    {
        "direction": "CHI",
        "nv_code": "NV1",
        "nv_description": "Chi mua quà Tết cho đoàn viên",
        "operation_type_code": "1.0",
        "operation_type_name": "Chi TCCĐ (theo mục chi) bằng tiền mặt",
        "fund_source_code": "11.0",
        "fund_source_name": "Đoàn phí CĐ",
        "budget_code": "31.03",
        "budget_name": "Thăm hỏi, trợ cấp",
        "sort_order": 10,
    },
    {
        "direction": "CHI",
        "nv_code": "NV2",
        "nv_description": "Thăm hỏi đoàn viên bị bệnh",
        "operation_type_code": "1.0",
        "operation_type_name": "Chi TCCĐ (theo mục chi) bằng tiền mặt",
        "fund_source_code": "11.0",
        "fund_source_name": "Đoàn phí CĐ",
        "budget_code": "31.03",
        "budget_name": "Thăm hỏi, trợ cấp",
        "sort_order": 20,
    },
    {
        "direction": "CHI",
        "nv_code": "NV3",
        "nv_description": "Chi tổ chức họp mặt ngày 08/03",
        "operation_type_code": "1.0",
        "operation_type_name": "Chi TCCĐ (theo mục chi) bằng tiền mặt",
        "fund_source_code": "12.0",
        "fund_source_name": "Kinh phí CĐ",
        "budget_code": "32.05",
        "budget_name": "Tổ chức hoạt động về giới, bình đẳng giới",
        "sort_order": 30,
    },
    {
        "direction": "CHI",
        "nv_code": "NV4",
        "nv_description": "Chi phụ cấp BCH, kế toán, thủ quỹ",
        "operation_type_code": "1.0",
        "operation_type_name": "Chi TCCĐ (theo mục chi) bằng tiền mặt",
        "fund_source_code": "11.0",
        "fund_source_name": "Đoàn phí CĐ",
        "budget_code": "34.02",
        "budget_name": "Chi phụ cấp cán bộ công đoàn",
        "sort_order": 40,
    },
    {
        "direction": "CHI",
        "nv_code": "NV5",
        "nv_description": "Nộp 30% đoàn phí lên cấp trên",
        "operation_type_code": "2.0",
        "operation_type_name": "Nộp đoàn phí",
        "fund_source_code": "11.0",
        "fund_source_name": "Đoàn phí CĐ",
        "budget_code": "39.0",
        "budget_name": "Tài chính",
        "sort_order": 50,
    },
    {
        "direction": "CHI",
        "nv_code": "NV6",
        "nv_description": "Phát sinh chi phí SMS, chuyển khoản",
        "operation_type_code": "1.0",
        "operation_type_name": "Chi TCCĐ (theo mục chi) bằng tiền mặt",
        "fund_source_code": "11.0",
        "fund_source_name": "Đoàn phí CĐ",
        "budget_code": "33.01",
        "budget_name": "TT DV công cộng",
        "sort_order": 60,
    },
    {
        "direction": "THU",
        "nv_code": "NV1",
        "nv_description": "Thu đoàn phí công đoàn",
        "operation_type_code": "1.0",
        "operation_type_name": "Thu TCCĐ",
        "fund_source_code": "11.0",
        "fund_source_name": "Đoàn phí CĐ",
        "budget_code": "22.01",
        "budget_name": "Thu đoàn phí khu vực SXKD",
        "sort_order": 10,
    },
    {
        "direction": "THU",
        "nv_code": "NV2",
        "nv_description": "Rút tiền từ ngân hàng nhập quỹ",
        "operation_type_code": "3.0",
        "operation_type_name": "Rút tiền gửi về nhập quỹ",
        "fund_source_code": "21.0/12.0",
        "fund_source_name": "Quỹ hoạt động thường xuyên / Kinh phí công đoàn",
        "budget_code": "",
        "budget_name": "",
        "sort_order": 20,
    },
    {
        "direction": "THU",
        "nv_code": "NV3",
        "nv_description": "Nhận tiền lãi",
        "operation_type_code": "1.0",
        "operation_type_name": "Thu TCCĐ",
        "fund_source_code": "142.0",
        "fund_source_name": "Lãi tiền gửi",
        "budget_code": "25.02",
        "budget_name": "Thu khác tại đơn vị",
        "sort_order": 30,
    },
    {
        "direction": "THU",
        "nv_code": "NV4",
        "nv_description": "Nhận kinh phí cấp trên",
        "operation_type_code": "2.0",
        "operation_type_name": "Nhận KP cấp trên bằng tiền gửi",
        "fund_source_code": "12.0",
        "fund_source_name": "Kinh phí CĐ",
        "budget_code": "28.01",
        "budget_name": "KPCĐ cấp trên theo phân phối",
        "sort_order": 40,
    },
]

DEFAULT_SOLAR_UNION_EVENTS = [
    ("Tết dương lịch", 1, 1),
    ("Quốc tế phụ nữ 8/3", 3, 8),
    ("Lễ 30/4 và 1/5", 4, 30),
    ("Quốc khánh 2/9", 9, 2),
    ("Phụ nữ Việt Nam 20/10", 10, 20),
    ("Quốc tế nam giới 19/11", 11, 19),
]

DEFAULT_LUNAR_UNION_EVENTS = [
    "Giỗ Tổ Hùng Vương 10/3 AL",
    "Tết Trung Thu",
    "Tết âm lịch",
]


def _to_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _safe_year(value):
    raw = str(value or "").strip()
    if raw.isdigit() and len(raw) == 4:
        return int(raw)
    return datetime.now().year


def _safe_month_key(value, year):
    raw = str(value or "").strip()
    if len(raw) == 7 and raw[4] == "-" and raw[:4].isdigit() and raw[5:].isdigit():
        return raw
    month = datetime.now().month
    return f"{year}-{month:02d}"


def _parse_date(value):
    raw = str(value or "").strip()
    if not raw:
        return None
    return datetime.strptime(raw, "%Y-%m-%d").date()


def _quarter_for_date(day_value):
    return ((day_value.month - 1) // 3) + 1


def _form_int(field_name):
    raw = str(request.form.get(field_name) or "").strip()
    if raw.isdigit():
        return int(raw)
    return None


def _redirect_period(endpoint, year, month_key, **kwargs):
    return redirect(url_for(endpoint, year=year, month_key=month_key, **kwargs))


def _default_union_holiday_payloads(year):
    rows = []
    for name, month, day in DEFAULT_SOLAR_UNION_EVENTS:
        event_date = date(year, month, day)
        rows.append(
            {
                "event_name": name,
                "event_date": event_date,
                "quarter": _quarter_for_date(event_date),
                "planned_amount": 0,
                "is_default": True,
                "notes": "Mặc định theo năm",
            }
        )

    for name in DEFAULT_LUNAR_UNION_EVENTS:
        rows.append(
            {
                "event_name": name,
                "event_date": None,
                "quarter": None,
                "planned_amount": 0,
                "is_default": True,
                "notes": "Cần cập nhật ngày âm lịch theo từng năm",
            }
        )

    return rows


def _ensure_union_year_seed(year):
    config = UnionYearConfig.query.filter_by(year=year).first()
    if not config:
        config = UnionYearConfig(
            year=year,
            opening_bank_balance=0,
            opening_cash_balance=0,
            notes="Khởi tạo mặc định",
        )
        db.session.add(config)

    has_hd = UnionHdRule.query.filter_by(year=year).first() is not None
    if not has_hd:
        for item in DEFAULT_UNION_HD_RULES:
            db.session.add(UnionHdRule(year=year, **item))

    has_holiday = UnionHolidayEvent.query.filter_by(year=year).first() is not None
    if not has_holiday:
        for item in _default_union_holiday_payloads(year):
            db.session.add(UnionHolidayEvent(year=year, **item))

    db.session.commit()


def _build_ledger_view(entries, opening_balance):
    running = float(opening_balance)
    rows = []
    quarter_summary = {q: {"in": 0.0, "out": 0.0, "net": 0.0} for q in (1, 2, 3, 4)}
    total_in = 0.0
    total_out = 0.0

    sorted_entries = sorted(
        entries,
        key=lambda row: (
            int(row.quarter or 0),
            row.event_date.isoformat() if row.event_date else "",
            row.id,
        ),
    )

    for row in sorted_entries:
        amount_in = _to_float(row.amount_in)
        amount_out = _to_float(row.amount_out)
        running += amount_in - amount_out

        quarter = int(row.quarter or (_quarter_for_date(row.event_date) if row.event_date else 0) or 0)
        if quarter in quarter_summary:
            quarter_summary[quarter]["in"] += amount_in
            quarter_summary[quarter]["out"] += amount_out
            quarter_summary[quarter]["net"] += amount_in - amount_out

        total_in += amount_in
        total_out += amount_out

        rows.append(
            {
                "entry": row,
                "amount_in": amount_in,
                "amount_out": amount_out,
                "running_balance": running,
                "quarter": quarter,
            }
        )

    return {
        "rows": rows,
        "quarter_summary": quarter_summary,
        "total_in": total_in,
        "total_out": total_out,
        "closing_balance": running,
    }


def _month_employee_reference(month_key):
    start_date, end_date = parse_month_key(month_key)

    active_total = Employee.query.filter_by(is_active=True).count()

    schedule_rows = (
        db.session.query(WorkSchedule.employee_id)
        .filter(WorkSchedule.work_date >= start_date, WorkSchedule.work_date <= end_date)
        .distinct()
        .all()
    )
    scheduled_ids = {row[0] for row in schedule_rows}

    worked_rows = (
        db.session.query(AttendanceDetail.employee_id)
        .filter(
            AttendanceDetail.month_key == month_key,
            AttendanceDetail.status_code.notin_(["OFF", "N"]),
        )
        .distinct()
        .all()
    )
    worked_ids = {row[0] for row in worked_rows}

    top_rows = []
    if scheduled_ids:
        top_rows = (
            Employee.query.filter(Employee.id.in_(scheduled_ids))
            .order_by(Employee.employee_code.asc())
            .limit(12)
            .all()
        )

    return {
        "active_total": active_total,
        "scheduled_total": len(scheduled_ids),
        "worked_total": len(worked_ids),
        "off_or_leave_total": max(len(scheduled_ids) - len(worked_ids), 0),
        "sample_rows": top_rows,
    }


def _employee_pool_for_event(month_key, scope):
    normalized_scope = str(scope or "ACTIVE").strip().upper()

    if normalized_scope == "SCHEDULED":
        start_date, end_date = parse_month_key(month_key)
        schedule_rows = (
            db.session.query(WorkSchedule.employee_id)
            .filter(WorkSchedule.work_date >= start_date, WorkSchedule.work_date <= end_date)
            .distinct()
            .all()
        )
        scheduled_ids = [row[0] for row in schedule_rows if row[0]]
        if scheduled_ids:
            scheduled_employees = (
                Employee.query.filter(Employee.id.in_(scheduled_ids), Employee.is_active.is_(True))
                .order_by(Employee.employee_code.asc())
                .all()
            )
            if scheduled_employees:
                return scheduled_employees, "SCHEDULED"

    active_employees = Employee.query.filter_by(is_active=True).order_by(Employee.employee_code.asc()).all()
    return active_employees, "ACTIVE"


def _rebuild_holiday_recipients(event_row, month_key, scope, default_amount):
    old_rows = UnionHolidayRecipient.query.filter_by(holiday_event_id=event_row.id).all()
    for row in old_rows:
        db.session.delete(row)

    employees, used_scope = _employee_pool_for_event(month_key, scope)
    for index, employee in enumerate(employees, start=1):
        db.session.add(
            UnionHolidayRecipient(
                holiday_event_id=event_row.id,
                employee_id=employee.id,
                employee_code=employee.employee_code,
                full_name=employee.full_name,
                gender=employee.gender,
                amount=default_amount,
                notes=None,
                sort_order=index,
            )
        )

    event_row.planned_amount = default_amount
    return len(employees), used_scope


def _build_union_year_views(year):
    config = UnionYearConfig.query.filter_by(year=year).first()

    ledger_rows = UnionLedgerEntry.query.filter_by(year=year).order_by(
        UnionLedgerEntry.source.asc(),
        UnionLedgerEntry.quarter.asc(),
        UnionLedgerEntry.event_date.asc(),
        UnionLedgerEntry.id.asc(),
    ).all()
    bank_entries = [row for row in ledger_rows if (row.source or "").upper() == "BANK"]
    cash_entries = [row for row in ledger_rows if (row.source or "").upper() == "CASH"]

    opening_bank = _to_float(config.opening_bank_balance if config else 0)
    opening_cash = _to_float(config.opening_cash_balance if config else 0)

    bank_view = _build_ledger_view(bank_entries, opening_bank)
    cash_view = _build_ledger_view(cash_entries, opening_cash)

    return config, bank_view, cash_view


def _holiday_event_rows_with_stats(year):
    events = UnionHolidayEvent.query.filter_by(year=year).all()
    events = sorted(
        events,
        key=lambda row: (
            int(row.quarter or 99),
            row.event_date.isoformat() if row.event_date else "9999-12-31",
            (row.event_name or "").lower(),
        ),
    )

    stat_rows = (
        db.session.query(
            UnionHolidayRecipient.holiday_event_id,
            func.count(UnionHolidayRecipient.id),
            func.coalesce(func.sum(UnionHolidayRecipient.amount), 0),
        )
        .join(UnionHolidayEvent, UnionHolidayEvent.id == UnionHolidayRecipient.holiday_event_id)
        .filter(UnionHolidayEvent.year == year)
        .group_by(UnionHolidayRecipient.holiday_event_id)
        .all()
    )

    stat_map = {
        row[0]: {
            "recipient_count": int(row[1] or 0),
            "total_amount": row[2] or 0,
        }
        for row in stat_rows
    }

    return [
        {
            "event": event,
            "recipient_count": stat_map.get(event.id, {}).get("recipient_count", 0),
            "total_amount": stat_map.get(event.id, {}).get("total_amount", 0),
        }
        for event in events
    ]


def _save_year_config(year, actor):
    row = UnionYearConfig.query.filter_by(year=year).first()
    before_data = row.to_dict() if row else None
    if not row:
        row = UnionYearConfig(year=year)
        db.session.add(row)

    row.opening_bank_balance = _to_float(request.form.get("opening_bank_balance"), 0)
    row.opening_cash_balance = _to_float(request.form.get("opening_cash_balance"), 0)
    row.notes = (request.form.get("notes") or "").strip() or None

    db.session.flush()
    log_action(
        "union_year_configs",
        str(row.id),
        "UPDATE" if before_data else "INSERT",
        changed_by=actor,
        before_data=before_data,
        after_data=row.to_dict(),
        notes=f"Lưu cấu hình Công đoàn năm {year}",
    )
    db.session.commit()
    flash("Đã lưu cấu hình năm.", "success")
    return True


def _add_ledger_entry(year, source, actor):
    source = str(source or "").upper()
    if source not in {"BANK", "CASH"}:
        flash("Nguồn sổ không hợp lệ.", "error")
        return False

    event_date = _parse_date(request.form.get("event_date"))
    if not event_date or event_date.year != year:
        flash(f"Ngày phát sinh phải thuộc năm {year}.", "error")
        return False

    description = (request.form.get("description") or "").strip()
    if not description:
        flash("Bạn cần nhập nội dung giao dịch.", "error")
        return False

    amount_in = _to_float(request.form.get("amount_in"), 0)
    amount_out = _to_float(request.form.get("amount_out"), 0)
    if amount_in <= 0 and amount_out <= 0:
        flash("Bạn cần nhập số tiền thu hoặc chi.", "error")
        return False

    quarter_value = request.form.get("quarter")
    quarter = int(quarter_value) if str(quarter_value or "").isdigit() else _quarter_for_date(event_date)
    if quarter not in {1, 2, 3, 4}:
        quarter = _quarter_for_date(event_date)

    row = UnionLedgerEntry(
        year=year,
        source=source,
        quarter=quarter,
        event_date=event_date,
        voucher_code=(request.form.get("voucher_code") or "").strip() or None,
        receipt_code=(request.form.get("receipt_code") or "").strip() or None,
        payment_code=(request.form.get("payment_code") or "").strip() or None,
        description=description,
        amount_in=amount_in,
        amount_out=amount_out,
        notes=(request.form.get("entry_notes") or "").strip() or None,
    )
    db.session.add(row)
    db.session.flush()

    source_label = "chuyển khoản" if source == "BANK" else "tiền mặt"
    log_action(
        "union_ledger_entries",
        str(row.id),
        "INSERT",
        changed_by=actor,
        after_data=row.to_dict(),
        notes=f"Thêm giao dịch sổ {source_label} năm {year}",
    )
    db.session.commit()
    flash("Đã thêm giao dịch.", "success")
    return True


def _delete_ledger_entry(year, source, actor):
    source = str(source or "").upper()
    row_id = _form_int("entry_id")
    if not row_id:
        flash("Không tìm thấy giao dịch để xóa.", "error")
        return False

    row = UnionLedgerEntry.query.get(row_id)
    if not row or row.year != year or (row.source or "").upper() != source:
        flash("Giao dịch không tồn tại hoặc không thuộc đúng sổ.", "error")
        return False

    before_data = row.to_dict()
    db.session.delete(row)

    source_label = "chuyển khoản" if source == "BANK" else "tiền mặt"
    log_action(
        "union_ledger_entries",
        str(row.id),
        "DELETE",
        changed_by=actor,
        before_data=before_data,
        notes=f"Xóa giao dịch sổ {source_label} năm {year}",
    )
    db.session.commit()
    flash("Đã xóa giao dịch.", "success")
    return True


def _create_holiday_event(year, month_key, actor):
    event_name = (request.form.get("event_name") or "").strip()
    if not event_name:
        flash("Bạn cần nhập tên sự kiện.", "error")
        return None

    default_amount = _to_float(request.form.get("default_amount"), 0)
    if default_amount < 0:
        flash("Số tiền mặc định phải lớn hơn hoặc bằng 0.", "error")
        return None

    existing = UnionHolidayEvent.query.filter_by(year=year, event_name=event_name).first()
    if existing:
        flash("Sự kiện này đã tồn tại trong năm, vui lòng đặt tên khác.", "error")
        return None

    event_date = _parse_date(request.form.get("event_date"))
    if event_date and event_date.year != year:
        flash(f"Ngày sự kiện phải thuộc năm {year}.", "error")
        return None

    quarter = _quarter_for_date(event_date) if event_date else None
    roster_scope = (request.form.get("roster_scope") or "ACTIVE").strip().upper()

    row = UnionHolidayEvent(
        year=year,
        event_name=event_name,
        event_date=event_date,
        quarter=quarter,
        planned_amount=default_amount,
        is_default=False,
        notes=(request.form.get("holiday_notes") or "").strip() or None,
    )
    db.session.add(row)
    db.session.flush()

    recipient_count, used_scope = _rebuild_holiday_recipients(
        event_row=row,
        month_key=month_key,
        scope=roster_scope,
        default_amount=default_amount,
    )

    db.session.flush()
    log_action(
        "union_holiday_events",
        str(row.id),
        "INSERT",
        changed_by=actor,
        after_data=row.to_dict(),
        notes=f"Tạo sự kiện lễ và sinh danh sách nhận tiền ({recipient_count} nhân viên)",
    )
    db.session.commit()

    if roster_scope == "SCHEDULED" and used_scope != "SCHEDULED":
        flash(
            "Đã tạo sự kiện. Không có nhân viên có lịch trong tháng nên hệ thống dùng danh sách nhân viên đang hoạt động.",
            "warning",
        )
    else:
        flash(f"Đã tạo sự kiện và sinh danh sách {recipient_count} nhân viên.", "success")

    return row.id


def _delete_holiday_event(year, actor):
    event_id = _form_int("holiday_id")
    if not event_id:
        flash("Không tìm thấy sự kiện để xóa.", "error")
        return False

    row = UnionHolidayEvent.query.get(event_id)
    if not row or row.year != year:
        flash("Sự kiện không tồn tại.", "error")
        return False

    before_data = row.to_dict()
    recipient_rows = UnionHolidayRecipient.query.filter_by(holiday_event_id=row.id).all()
    recipient_count = len(recipient_rows)

    for recipient in recipient_rows:
        db.session.delete(recipient)
    db.session.delete(row)

    log_action(
        "union_holiday_events",
        str(row.id),
        "DELETE",
        changed_by=actor,
        before_data=before_data,
        notes=f"Xóa sự kiện lễ và {recipient_count} dòng danh sách nhận tiền",
    )
    db.session.commit()
    flash("Đã xóa sự kiện lễ.", "success")
    return True


def _fill_default_holidays(year):
    existing_names = {row.event_name for row in UnionHolidayEvent.query.filter_by(year=year).all()}
    inserted = 0
    for item in _default_union_holiday_payloads(year):
        if item["event_name"] in existing_names:
            continue
        db.session.add(UnionHolidayEvent(year=year, **item))
        inserted += 1
    db.session.commit()
    return inserted


def _update_event_metadata(event_row, year, actor):
    event_name = (request.form.get("event_name") or "").strip()
    if not event_name:
        flash("Tên sự kiện không được để trống.", "error")
        return False

    duplicated = UnionHolidayEvent.query.filter_by(year=year, event_name=event_name).first()
    if duplicated and duplicated.id != event_row.id:
        flash("Tên sự kiện đã tồn tại, vui lòng chọn tên khác.", "error")
        return False

    event_date = _parse_date(request.form.get("event_date"))
    if event_date and event_date.year != year:
        flash(f"Ngày sự kiện phải thuộc năm {year}.", "error")
        return False

    default_amount = _to_float(request.form.get("default_amount"), event_row.planned_amount)
    if default_amount < 0:
        flash("Số tiền mặc định phải lớn hơn hoặc bằng 0.", "error")
        return False

    before_data = event_row.to_dict()

    event_row.event_name = event_name
    event_row.event_date = event_date
    event_row.quarter = _quarter_for_date(event_date) if event_date else None
    event_row.planned_amount = default_amount
    event_row.notes = (request.form.get("holiday_notes") or "").strip() or None

    db.session.flush()
    log_action(
        "union_holiday_events",
        str(event_row.id),
        "UPDATE",
        changed_by=actor,
        before_data=before_data,
        after_data=event_row.to_dict(),
        notes="Cập nhật thông tin sự kiện lễ",
    )
    db.session.commit()
    flash("Đã cập nhật thông tin sự kiện.", "success")
    return True


def _regenerate_event_recipients(event_row, month_key, actor):
    default_amount = _to_float(request.form.get("default_amount"), event_row.planned_amount)
    if default_amount < 0:
        flash("Số tiền mặc định phải lớn hơn hoặc bằng 0.", "error")
        return False

    scope = (request.form.get("roster_scope") or "ACTIVE").strip().upper()
    old_count = UnionHolidayRecipient.query.filter_by(holiday_event_id=event_row.id).count()
    before_data = event_row.to_dict()

    new_count, used_scope = _rebuild_holiday_recipients(
        event_row=event_row,
        month_key=month_key,
        scope=scope,
        default_amount=default_amount,
    )

    db.session.flush()
    log_action(
        "union_holiday_events",
        str(event_row.id),
        "UPDATE",
        changed_by=actor,
        before_data=before_data,
        after_data=event_row.to_dict(),
        notes=f"Làm mới danh sách nhận tiền: {old_count} -> {new_count}",
    )
    db.session.commit()

    if scope == "SCHEDULED" and used_scope != "SCHEDULED":
        flash(
            "Đã làm mới danh sách. Không có nhân viên có lịch trong tháng nên hệ thống dùng danh sách nhân viên đang hoạt động.",
            "warning",
        )
    else:
        flash(f"Đã làm mới danh sách nhân viên ({new_count} người).", "success")

    return True


def _add_event_recipient(event_row, actor):
    employee_id = _form_int("employee_id")
    if not employee_id:
        flash("Bạn cần chọn nhân viên để thêm.", "error")
        return False

    employee = Employee.query.get(employee_id)
    if not employee:
        flash("Nhân viên không tồn tại.", "error")
        return False

    duplicated = UnionHolidayRecipient.query.filter_by(
        holiday_event_id=event_row.id,
        employee_id=employee.id,
    ).first()
    if duplicated:
        flash("Nhân viên này đã có trong danh sách nhận tiền.", "warning")
        return False

    amount = _to_float(request.form.get("amount"), event_row.planned_amount)
    if amount < 0:
        flash("Số tiền phải lớn hơn hoặc bằng 0.", "error")
        return False

    max_sort = (
        db.session.query(func.coalesce(func.max(UnionHolidayRecipient.sort_order), 0))
        .filter(UnionHolidayRecipient.holiday_event_id == event_row.id)
        .scalar()
    )

    row = UnionHolidayRecipient(
        holiday_event_id=event_row.id,
        employee_id=employee.id,
        employee_code=employee.employee_code,
        full_name=employee.full_name,
        gender=employee.gender,
        amount=amount,
        notes=(request.form.get("recipient_notes") or "").strip() or None,
        sort_order=int(max_sort or 0) + 1,
    )
    db.session.add(row)
    db.session.flush()

    log_action(
        "union_holiday_recipients",
        str(row.id),
        "INSERT",
        changed_by=actor,
        after_data=row.to_dict(),
        notes=f"Thêm nhân viên vào sự kiện lễ #{event_row.id}",
    )
    db.session.commit()
    flash("Đã thêm nhân viên vào danh sách nhận tiền.", "success")
    return True


def _update_event_recipient_amount(event_row, actor):
    recipient_id = _form_int("recipient_id")
    if not recipient_id:
        flash("Không tìm thấy dòng nhân viên cần sửa.", "error")
        return False

    row = UnionHolidayRecipient.query.filter_by(id=recipient_id, holiday_event_id=event_row.id).first()
    if not row:
        flash("Dòng nhân viên không tồn tại.", "error")
        return False

    amount = _to_float(request.form.get("amount"), row.amount)
    if amount < 0:
        flash("Số tiền phải lớn hơn hoặc bằng 0.", "error")
        return False

    before_data = row.to_dict()

    row.amount = amount
    row.notes = (request.form.get("recipient_notes") or "").strip() or None

    db.session.flush()
    log_action(
        "union_holiday_recipients",
        str(row.id),
        "UPDATE",
        changed_by=actor,
        before_data=before_data,
        after_data=row.to_dict(),
        notes=f"Cập nhật số tiền nhân viên trong sự kiện lễ #{event_row.id}",
    )
    db.session.commit()
    flash("Đã cập nhật số tiền nhân viên.", "success")
    return True


def _delete_event_recipient(event_row, actor):
    recipient_id = _form_int("recipient_id")
    if not recipient_id:
        flash("Không tìm thấy dòng nhân viên cần xóa.", "error")
        return False

    row = UnionHolidayRecipient.query.filter_by(id=recipient_id, holiday_event_id=event_row.id).first()
    if not row:
        flash("Dòng nhân viên không tồn tại.", "error")
        return False

    before_data = row.to_dict()
    db.session.delete(row)

    log_action(
        "union_holiday_recipients",
        str(row.id),
        "DELETE",
        changed_by=actor,
        before_data=before_data,
        notes=f"Xóa nhân viên khỏi sự kiện lễ #{event_row.id}",
    )
    db.session.commit()
    flash("Đã xóa nhân viên khỏi danh sách nhận tiền.", "success")
    return True


def _apply_amount_for_all(event_row, actor):
    amount = _to_float(request.form.get("amount"), event_row.planned_amount)
    if amount < 0:
        flash("Số tiền phải lớn hơn hoặc bằng 0.", "error")
        return False

    rows = UnionHolidayRecipient.query.filter_by(holiday_event_id=event_row.id).all()
    if not rows:
        flash("Danh sách nhân viên đang trống.", "warning")
        return False

    before_data = event_row.to_dict()
    for row in rows:
        row.amount = amount

    event_row.planned_amount = amount

    db.session.flush()
    log_action(
        "union_holiday_events",
        str(event_row.id),
        "UPDATE",
        changed_by=actor,
        before_data=before_data,
        after_data=event_row.to_dict(),
        notes=f"Áp dụng đồng loạt mức {amount:,.0f} cho {len(rows)} nhân viên",
    )
    db.session.commit()
    flash("Đã áp dụng đồng loạt số tiền cho toàn bộ nhân viên trong sự kiện.", "success")
    return True


def register_union_routes(app):
    def _require_admin():
        if session.get("is_admin"):
            return None
        flash("Bạn không có quyền truy cập chức năng Công đoàn.", "error")
        return redirect(url_for("dashboard"))

    @app.route("/union", methods=["GET", "POST"])
    def union_index():
        blocked = _require_admin()
        if blocked:
            return blocked

        year = _safe_year(request.values.get("year"))
        month_key = _safe_month_key(request.values.get("month_key"), year)
        actor = session.get("username") or "admin"

        _ensure_union_year_seed(year)

        if request.method == "POST":
            action = (request.form.get("action") or "").strip().lower()
            if action == "save_year_config":
                _save_year_config(year, actor)
            return _redirect_period("union_index", year, month_key)

        config, bank_view, cash_view = _build_union_year_views(year)
        month_ref = _month_employee_reference(month_key)
        holiday_rows = _holiday_event_rows_with_stats(year)
        total_holiday_budget = sum(_to_float(item["total_amount"], 0) for item in holiday_rows)

        return render_template(
            "union_home.html",
            title="Công đoàn - Tổng quan",
            active_tab="home",
            year=year,
            month_key=month_key,
            config_row=config,
            bank_view=bank_view,
            cash_view=cash_view,
            combined_closing=bank_view["closing_balance"] + cash_view["closing_balance"],
            month_ref=month_ref,
            holiday_rows=holiday_rows,
            total_holiday_budget=total_holiday_budget,
        )

    @app.route("/union/bank", methods=["GET", "POST"])
    def union_bank_page():
        blocked = _require_admin()
        if blocked:
            return blocked

        year = _safe_year(request.values.get("year"))
        month_key = _safe_month_key(request.values.get("month_key"), year)
        actor = session.get("username") or "admin"

        _ensure_union_year_seed(year)

        if request.method == "POST":
            action = (request.form.get("action") or "").strip().lower()
            if action == "save_year_config":
                _save_year_config(year, actor)
            elif action == "add_bank_ledger":
                _add_ledger_entry(year, "BANK", actor)
            elif action == "delete_ledger":
                _delete_ledger_entry(year, "BANK", actor)
            return _redirect_period("union_bank_page", year, month_key)

        config, bank_view, _ = _build_union_year_views(year)

        return render_template(
            "union_bank.html",
            title="Công đoàn - Chuyển khoản",
            active_tab="bank",
            year=year,
            month_key=month_key,
            config_row=config,
            bank_view=bank_view,
        )

    @app.route("/union/cash", methods=["GET", "POST"])
    def union_cash_page():
        blocked = _require_admin()
        if blocked:
            return blocked

        year = _safe_year(request.values.get("year"))
        month_key = _safe_month_key(request.values.get("month_key"), year)
        actor = session.get("username") or "admin"

        _ensure_union_year_seed(year)

        if request.method == "POST":
            action = (request.form.get("action") or "").strip().lower()
            if action == "save_year_config":
                _save_year_config(year, actor)
            elif action == "add_cash_ledger":
                _add_ledger_entry(year, "CASH", actor)
            elif action == "delete_ledger":
                _delete_ledger_entry(year, "CASH", actor)
            return _redirect_period("union_cash_page", year, month_key)

        config, _, cash_view = _build_union_year_views(year)

        return render_template(
            "union_cash.html",
            title="Công đoàn - Tiền mặt",
            active_tab="cash",
            year=year,
            month_key=month_key,
            config_row=config,
            cash_view=cash_view,
        )

    @app.route("/union/events", methods=["GET", "POST"])
    def union_events_page():
        blocked = _require_admin()
        if blocked:
            return blocked

        year = _safe_year(request.values.get("year"))
        month_key = _safe_month_key(request.values.get("month_key"), year)
        actor = session.get("username") or "admin"

        _ensure_union_year_seed(year)

        if request.method == "POST":
            action = (request.form.get("action") or "").strip().lower()

            if action == "create_holiday_event":
                event_id = _create_holiday_event(year, month_key, actor)
                if event_id:
                    return _redirect_period("union_event_detail", year, month_key, event_id=event_id)
            elif action == "delete_holiday_event":
                _delete_holiday_event(year, actor)
            elif action == "fill_default_holidays":
                inserted = _fill_default_holidays(year)
                flash(f"Đã bổ sung {inserted} sự kiện mặc định còn thiếu.", "success")

            return _redirect_period("union_events_page", year, month_key)

        holiday_rows = _holiday_event_rows_with_stats(year)
        month_ref = _month_employee_reference(month_key)

        return render_template(
            "union_events.html",
            title="Công đoàn - Sự kiện lễ",
            active_tab="events",
            year=year,
            month_key=month_key,
            holiday_rows=holiday_rows,
            month_ref=month_ref,
        )

    @app.route("/union/events/<int:event_id>", methods=["GET", "POST"])
    def union_event_detail(event_id):
        blocked = _require_admin()
        if blocked:
            return blocked

        event_row = UnionHolidayEvent.query.get(event_id)
        if not event_row:
            flash("Không tìm thấy sự kiện lễ.", "error")
            return redirect(url_for("union_events_page"))

        year = _safe_year(request.values.get("year") or event_row.year)
        if event_row.year != year:
            year = event_row.year

        month_key = _safe_month_key(request.values.get("month_key"), year)
        actor = session.get("username") or "admin"

        _ensure_union_year_seed(year)

        if request.method == "POST":
            action = (request.form.get("action") or "").strip().lower()

            if action == "update_event_meta":
                _update_event_metadata(event_row, year, actor)
            elif action == "regenerate_recipients":
                _regenerate_event_recipients(event_row, month_key, actor)
            elif action == "add_recipient":
                _add_event_recipient(event_row, actor)
            elif action == "update_recipient_amount":
                _update_event_recipient_amount(event_row, actor)
            elif action == "delete_recipient":
                _delete_event_recipient(event_row, actor)
            elif action == "apply_amount_all":
                _apply_amount_for_all(event_row, actor)

            return _redirect_period("union_event_detail", year, month_key, event_id=event_row.id)

        recipients = UnionHolidayRecipient.query.filter_by(holiday_event_id=event_row.id).order_by(
            UnionHolidayRecipient.sort_order.asc(),
            UnionHolidayRecipient.employee_code.asc(),
            UnionHolidayRecipient.id.asc(),
        ).all()

        total_amount = sum(_to_float(row.amount, 0) for row in recipients)

        recipient_ids = {row.employee_id for row in recipients if row.employee_id}
        employee_query = Employee.query.filter(Employee.is_active.is_(True))
        if recipient_ids:
            employee_query = employee_query.filter(~Employee.id.in_(recipient_ids))
        available_employees = employee_query.order_by(Employee.employee_code.asc()).all()

        return render_template(
            "union_event_detail.html",
            title=f"Công đoàn - {event_row.event_name}",
            active_tab="events",
            year=year,
            month_key=month_key,
            event_row=event_row,
            recipients=recipients,
            total_amount=total_amount,
            available_employees=available_employees,
        )

    # ── Excel exports ─────────────────────────────────────────────────

    def _xl_header(ws, cols):
        hdr_fill = PatternFill("solid", fgColor="1E3A5F")
        hdr_font = Font(color="FFFFFF", bold=True, size=10)
        ws.append(cols)
        for cell in ws[1]:
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="center")

    def _xl_response(wb, filename):
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, download_name=filename, as_attachment=True,
                         mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    @app.route("/union/bank/export.xlsx")
    def union_bank_export():
        blocked = _require_admin()
        if blocked:
            return blocked
        year = _safe_year(request.args.get("year"))
        config, bank_view, _ = _build_union_year_views(year)
        opening = _to_float(config.opening_bank_balance if config else 0)

        wb = Workbook()
        ws = wb.active
        ws.title = f"CK {year}"
        _xl_header(ws, ["STT", "Ngày", "Quý", "Số hiệu chứng từ", "Nội dung", "Tiền vào", "Tiền ra", "Còn lại", "Ghi chú"])
        ws.append(["", "", "", "Số dư đầu năm", "", opening, "", opening, ""])
        for i, item in enumerate(bank_view["rows"], 1):
            e = item["entry"]
            ws.append([i, str(e.event_date or ""), f"Q{item['quarter']}",
                       e.voucher_code or "", e.description,
                       float(item["amount_in"]), float(item["amount_out"]),
                       float(item["running_balance"]), e.notes or ""])
        ws.append(["", "", "", "TỔNG CỘNG", "",
                   bank_view["total_in"], bank_view["total_out"],
                   bank_view["closing_balance"], ""])
        for col in ["F", "G", "H"]:
            for row_idx in range(2, ws.max_row + 1):
                ws[f"{col}{row_idx}"].number_format = '#,##0'
        return _xl_response(wb, f"so_ck_{year}.xlsx")

    @app.route("/union/cash/export.xlsx")
    def union_cash_export():
        blocked = _require_admin()
        if blocked:
            return blocked
        year = _safe_year(request.args.get("year"))
        config, _, cash_view = _build_union_year_views(year)
        opening = _to_float(config.opening_cash_balance if config else 0)

        wb = Workbook()
        ws = wb.active
        ws.title = f"TM {year}"
        _xl_header(ws, ["STT", "Ngày", "Quý", "Phiếu thu", "Phiếu chi", "Nội dung", "Thu", "Chi", "Tồn quỹ", "Ghi chú"])
        ws.append(["", "", "", "", "", "Số dư đầu năm", opening, "", opening, ""])
        for i, item in enumerate(cash_view["rows"], 1):
            e = item["entry"]
            ws.append([i, str(e.event_date or ""), f"Q{item['quarter']}",
                       e.receipt_code or "", e.payment_code or "", e.description,
                       float(item["amount_in"]), float(item["amount_out"]),
                       float(item["running_balance"]), e.notes or ""])
        ws.append(["", "", "", "", "", "TỔNG CỘNG",
                   cash_view["total_in"], cash_view["total_out"],
                   cash_view["closing_balance"], ""])
        for col in ["G", "H", "I"]:
            for row_idx in range(2, ws.max_row + 1):
                ws[f"{col}{row_idx}"].number_format = '#,##0'
        return _xl_response(wb, f"so_tm_{year}.xlsx")

    @app.route("/union/events/<int:event_id>/export.xlsx")
    def union_event_export(event_id):
        blocked = _require_admin()
        if blocked:
            return blocked
        event_row = UnionHolidayEvent.query.get(event_id)
        if not event_row:
            flash("Không tìm thấy sự kiện.", "error")
            return redirect(url_for("union_events_page"))

        recipients = UnionHolidayRecipient.query.filter_by(holiday_event_id=event_id).order_by(
            UnionHolidayRecipient.sort_order.asc(),
            UnionHolidayRecipient.employee_code.asc(),
        ).all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Danh sách chi"
        ws["A1"] = f"SỰ KIỆN: {event_row.event_name}"
        ws["A1"].font = Font(bold=True, size=12)
        ws["A2"] = f"Ngày: {event_row.event_date or 'Chưa chốt'}  |  Đơn giá: {float(event_row.planned_amount or 0):,.0f}"
        ws.append([])
        _xl_header(ws, ["STT", "Mã NV", "Họ tên", "Giới tính", "Số tiền", "Ghi chú", "Ký nhận"])
        for i, row in enumerate(recipients, 1):
            ws.append([i, row.employee_code, row.full_name, row.gender or "",
                       float(row.amount or 0), row.notes or "", ""])
        total = sum(_to_float(r.amount, 0) for r in recipients)
        ws.append(["", "", "TỔNG CỘNG", "", total, "", ""])
        for row_idx in range(5, ws.max_row + 1):
            ws[f"E{row_idx}"].number_format = '#,##0'
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 24
        ws.column_dimensions["E"].width = 14
        ws.column_dimensions["G"].width = 16
        safe_name = "".join(c for c in event_row.event_name if c.isalnum() or c in " _-")[:30].strip()
        return _xl_response(wb, f"ds_chi_{safe_name}_{event_row.year}.xlsx")
