import io
from datetime import date, datetime

from flask import request, send_file, session
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import func
from sqlalchemy.orm import joinedload

from ..database import db
from ..models import (
    AdvancePayment,
    AttendanceDetail,
    Employee,
    MonthlySalary,
    MonthlyWorkdayConfig,
    PayrollMonthLock,
    PayrollPaymentStatus,
)
from .attendance import parse_month_key


def _to_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _employee_code_sort_key(employee_code):
    raw_code = str(employee_code or "").replace("'", "").strip()
    if raw_code.isdigit():
        return (0, int(raw_code))
    return (1, raw_code.lower())


def _is_paid_off_detail(detail_row):
    status_code = str(getattr(detail_row, "status_code", "") or "").strip().upper()
    if status_code != "OFF":
        return False
    return _to_float(getattr(detail_row, "paid_hours", 0), 0.0) > 0


def _safe_month_key(value):
    if not value:
        return date.today().strftime("%Y-%m")

    try:
        datetime.strptime(str(value).strip(), "%Y-%m")
        return str(value).strip()
    except ValueError:
        return date.today().strftime("%Y-%m")


def _resolve_company_work_days(month_key):
    config = MonthlyWorkdayConfig.query.filter_by(month_key=month_key).first()
    config_value = _to_float(config.company_work_days if config else None, 0)
    if config_value > 0:
        return config_value, config

    legacy_salary = MonthlySalary.query.filter_by(month_key=month_key).order_by(MonthlySalary.id.asc()).first()
    legacy_value = _to_float(legacy_salary.salary_coefficient if legacy_salary else None, 0)
    if legacy_value >= 10:
        return legacy_value, config

    return 26.0, config


def _query_locked_month_keys(month_keys=None):
    query = PayrollMonthLock.query.filter(PayrollMonthLock.is_locked.is_(True))

    if month_keys is not None:
        keys = [item for item in month_keys if item]
        if not keys:
            return set()
        query = query.filter(PayrollMonthLock.month_key.in_(keys))

    return {row.month_key for row in query.all() if row.month_key}


def _collect_salary_overview_data(args):
    month_key = _safe_month_key(args.get("month"))
    search_query = (args.get("q") or "").strip()
    search_scope = (args.get("scope") or "current").strip().lower()
    if search_scope not in {"current", "all"}:
        search_scope = "current"

    start_date, end_date = parse_month_key(month_key)
    period_1_end = date(start_date.year, start_date.month, 15)
    period_2_start = date(start_date.year, start_date.month, 16)

    company_work_days_current, _ = _resolve_company_work_days(month_key)
    if company_work_days_current <= 0:
        company_work_days_current = 26.0

    summary_map = {}
    if search_scope == "current":
        employees = Employee.query.filter(Employee.is_active.is_(True)).order_by(Employee.employee_code.asc()).all()
        for row in employees:
            summary_map[(month_key, row.id)] = {
                "month_key": month_key,
                "employee": row,
                "worked_days": 0.0,
                "paid_leave_days": 0.0,
                "unpaid_leave_days": 0.0,
                "overtime_hours": 0.0,
            }

    detail_query = AttendanceDetail.query.options(joinedload(AttendanceDetail.employee))
    if search_scope == "current":
        detail_query = detail_query.filter(AttendanceDetail.month_key == month_key)

    detail_rows = (
        detail_query.order_by(
            AttendanceDetail.month_key.desc(),
            AttendanceDetail.employee_id.asc(),
            AttendanceDetail.work_date.asc(),
        ).all()
    )

    def _apply_status(summary_obj, detail_row):
        normalized = str(getattr(detail_row, "status_code", "") or "").upper()
        if normalized == "P":
            summary_obj["paid_leave_days"] += 1.0
        elif normalized in {"S", "C"}:
            summary_obj["paid_leave_days"] += 0.5
            summary_obj["worked_days"] += 0.5
        elif normalized == "N":
            summary_obj["unpaid_leave_days"] += 1.0
        elif normalized == "OFF":
            if _is_paid_off_detail(detail_row):
                summary_obj["worked_days"] += 1.0
            return
        else:
            summary_obj["worked_days"] += 1.0

    for row in detail_rows:
        if not row.employee:
            continue

        item_month = (row.month_key or month_key).strip()
        summary_key = (item_month, row.employee_id)
        summary = summary_map.get(summary_key)
        if not summary:
            summary = {
                "month_key": item_month,
                "employee": row.employee,
                "worked_days": 0.0,
                "paid_leave_days": 0.0,
                "unpaid_leave_days": 0.0,
                "overtime_hours": 0.0,
            }
            summary_map[summary_key] = summary

        _apply_status(summary, row)
        summary["overtime_hours"] += _to_float(row.overtime_hours)

    summary_keys = list(summary_map.keys())
    employee_ids = sorted({employee_id for _, employee_id in summary_keys})
    summary_month_keys = sorted({item_month for item_month, _ in summary_keys if item_month})
    is_admin_user = bool(session.get("is_admin"))

    salary_by_key = {}
    if employee_ids and summary_month_keys:
        salary_rows = MonthlySalary.query.filter(
            MonthlySalary.employee_id.in_(employee_ids),
            MonthlySalary.month_key.in_(summary_month_keys),
        ).all()
        salary_by_key = {
            (row.month_key, row.employee_id): row
            for row in salary_rows
        }

    payment_status_by_key = {}
    if employee_ids and summary_month_keys:
        payment_rows = PayrollPaymentStatus.query.filter(
            PayrollPaymentStatus.employee_id.in_(employee_ids),
            PayrollPaymentStatus.month_key.in_(summary_month_keys),
        ).all()
        payment_status_by_key = {
            (row.month_key, row.employee_id): row
            for row in payment_rows
        }

    lock_month_keys = set(summary_month_keys)
    lock_month_keys.add(month_key)
    locked_month_keys = _query_locked_month_keys(lock_month_keys)

    advance_by_key = {}
    if employee_ids and summary_month_keys:
        advance_rows = (
            db.session.query(
                AdvancePayment.month_key,
                AdvancePayment.employee_id,
                func.coalesce(func.sum(AdvancePayment.amount), 0),
            )
            .filter(
                AdvancePayment.month_key.in_(summary_month_keys),
                AdvancePayment.employee_id.in_(employee_ids),
            )
            .group_by(AdvancePayment.month_key, AdvancePayment.employee_id)
            .all()
        )
        advance_by_key = {
            (item_month, employee_id): _to_float(total_amount)
            for item_month, employee_id, total_amount in advance_rows
        }

    workday_by_month = {month_key: company_work_days_current}
    if search_scope == "all" and summary_month_keys:
        workday_rows = MonthlyWorkdayConfig.query.filter(
            MonthlyWorkdayConfig.month_key.in_(summary_month_keys)
        ).all()
        workday_by_month = {
            row.month_key: _to_float(row.company_work_days, 0)
            for row in workday_rows
        }
        for item_month in summary_month_keys:
            if _to_float(workday_by_month.get(item_month), 0) <= 0:
                workday_by_month[item_month] = 26.0

    overview_rows = []
    for (item_month, employee_id), summary in summary_map.items():
        salary_row = salary_by_key.get((item_month, employee_id))
        monthly_wage = _to_float(salary_row.base_daily_wage if salary_row else None, 0)

        overtime_hours = round(summary["overtime_hours"], 2)
        overtime_day_units = int(overtime_hours // 8)
        overtime_remainder_hours = round(overtime_hours - (overtime_day_units * 8), 2)

        salary_day_units = (
            summary["worked_days"]
            + summary["paid_leave_days"]
            + (overtime_hours / 8.0)
        )

        company_work_days = _to_float(workday_by_month.get(item_month), 0)
        if company_work_days <= 0:
            legacy_value = _to_float(salary_row.salary_coefficient if salary_row else None, 0)
            company_work_days = legacy_value if legacy_value >= 10 else 26.0

        daily_rate = (monthly_wage / company_work_days) if company_work_days > 0 else 0.0
        salary_amount = round(daily_rate * salary_day_units, 2)
        advance_amount = round(_to_float(advance_by_key.get((item_month, employee_id), 0), 0), 2)
        payment_status = payment_status_by_key.get((item_month, employee_id))
        month_is_locked = item_month in locked_month_keys

        overview_rows.append(
            {
                "month_key": item_month,
                "employee": summary["employee"],
                "worked_days": round(summary["worked_days"], 2),
                "paid_leave_days": round(summary["paid_leave_days"], 2),
                "unpaid_leave_days": round(summary["unpaid_leave_days"], 2),
                "overtime_day_units": overtime_day_units,
                "overtime_remainder_hours": overtime_remainder_hours,
                "advance_amount": advance_amount,
                "salary_amount": salary_amount,
                "salary_received": bool(payment_status.salary_received) if payment_status else False,
                "meal_period_1_received": bool(payment_status.meal_period_1_received) if payment_status else False,
                "meal_period_2_received": bool(payment_status.meal_period_2_received) if payment_status else False,
                "month_is_locked": month_is_locked,
                "can_update_payment_status": is_admin_user or not month_is_locked,
            }
        )

    if search_scope == "all":
        overview_rows.sort(
            key=lambda item: _employee_code_sort_key(item["employee"].employee_code)
        )
        overview_rows.sort(key=lambda item: item["month_key"], reverse=True)
    else:
        overview_rows.sort(
            key=lambda item: _employee_code_sort_key(item["employee"].employee_code)
        )

    if search_query:
        search_text = search_query.lower()

        def _match_overview(item):
            values = [
                item["month_key"],
                item["employee"].employee_code,
                item["employee"].full_name,
                item["worked_days"],
                item["paid_leave_days"],
                item["unpaid_leave_days"],
                item["overtime_day_units"],
                item["overtime_remainder_hours"],
                item["advance_amount"],
                item["salary_amount"],
                item["salary_received"],
                item["meal_period_1_received"],
                item["meal_period_2_received"],
            ]
            return any(
                search_text in str(value).lower()
                for value in values
                if value is not None
            )

        overview_rows = [item for item in overview_rows if _match_overview(item)]

    return {
        "month_key": month_key,
        "search_query": search_query,
        "search_scope": search_scope,
        "overview_rows": overview_rows,
        "company_work_days": round(company_work_days_current, 2),
        "period_1_end": period_1_end,
        "period_2_start": period_2_start,
    }


def export_salary_overview_excel():
    overview_data = _collect_salary_overview_data(request.args)
    rows = overview_data["overview_rows"]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Bang luong"

    headers = [
        "STT",
        "Họ tên",
        "Số ngày làm",
        "Số ngày nghỉ có phép",
        "Số ngày nghỉ không phép",
        "Số ngày tăng ca",
        "Số giờ lẻ",
        "Số tiền ứng",
        "Tiền lương",
    ]
    sheet.append(headers)

    salary_bold_font = Font(bold=True)

    for index, item in enumerate(rows, start=1):
        sheet.append(
            [
                index,
                item["employee"].full_name,
                float(item["worked_days"]),
                float(item["paid_leave_days"]),
                float(item["unpaid_leave_days"]),
                int(item["overtime_day_units"]),
                float(item["overtime_remainder_hours"]),
                float(item["advance_amount"]),
                float(item["salary_amount"]),
            ]
        )

        amount_cell = sheet.cell(row=sheet.max_row, column=8)
        salary_cell = sheet.cell(row=sheet.max_row, column=9)
        amount_cell.number_format = "#,##0"
        salary_cell.number_format = "#,##0"
        salary_cell.font = salary_bold_font

    sheet.freeze_panes = "A2"

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)

    month_label = overview_data["month_key"].replace("-", "")
    scope_label = "tat_ca" if overview_data["search_scope"] == "all" else "thang"
    filename = f"bang_luong_{scope_label}_{month_label}.xlsx"

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=filename,
    )


def register_salary_overview_export(app):
    app.view_functions["export_salary_overview_excel"] = export_salary_overview_excel