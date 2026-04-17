import re
import unicodedata
import uuid
from datetime import date, datetime

from openpyxl import load_workbook

from ..database import db
from ..models import AttendanceDetail, Employee, OvertimeEntry, ShiftTemplate, WorkSchedule
from .attendance import month_key_for_date, parse_month_key
from .audit import log_action


EMPLOYEE_HEADER_PATTERN = re.compile(r"^ID\s*'?\s*(\d+)\s*-\s*(.+)$", re.IGNORECASE)
SHEET_MONTH_PATTERN = re.compile(r"(\d{1,2})\s*[/\-]\s*(\d{4})")
PREFERRED_SCHEDULE_SHEET_NAME = "LICH LAM"
EMPTY_SHIFT_MARKERS = {"", "-", "NONE", "NAN", "NULL"}


def _parse_employee_header(value):
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    matched = EMPLOYEE_HEADER_PATTERN.match(text)
    if not matched:
        return None

    employee_code = matched.group(1).strip()
    employee_name = matched.group(2).strip()
    return employee_code, employee_name


def _fold_text(value):
    if value is None:
        return ""

    text = str(value).strip().upper()
    if not text:
        return ""

    normalized = unicodedata.normalize("NFKD", text)
    return "".join(char for char in normalized if not unicodedata.combining(char))


def _coerce_date(value):
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()
    if not text:
        return None

    parsed = None
    for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"]:
        try:
            parsed = datetime.strptime(text, fmt)
            break
        except ValueError:
            continue

    return parsed.date() if parsed else None


def _normalize_shift_code(value):
    if value is None:
        return None

    text = str(value).strip().upper()
    if text in EMPTY_SHIFT_MARKERS:
        return None

    return text


def _is_blank_cell(value):
    if value is None:
        return True
    return isinstance(value, str) and not value.strip()


def _coerce_day_number(value):
    if value is None:
        return None

    day = None
    if isinstance(value, int):
        day = value
    elif isinstance(value, float):
        if not value.is_integer():
            return None
        day = int(value)
    else:
        text = str(value).strip()
        if not text or not text.isdigit():
            return None
        day = int(text)

    if 1 <= day <= 31:
        return day
    return None


def _coerce_employee_code(value):
    if value is None:
        return None

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()

    text = str(value).strip()
    if not text:
        return None

    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]

    return text


def _resolve_schedule_worksheet(workbook):
    for sheet_name in workbook.sheetnames:
        if _fold_text(sheet_name) == PREFERRED_SCHEDULE_SHEET_NAME:
            return workbook[sheet_name]
    return workbook[workbook.sheetnames[0]]


def _detect_grid_schedule_layout(worksheet):
    scan_limit = min(15, worksheet.max_row)

    for row in range(1, scan_limit + 1):
        if _fold_text(worksheet.cell(row, 2).value) != "MA NV":
            continue
        if _fold_text(worksheet.cell(row, 3).value) != "HO TEN":
            continue

        day_columns = []
        for column in range(4, worksheet.max_column + 1):
            day = _coerce_day_number(worksheet.cell(row, column).value)
            if day is None:
                continue
            day_columns.append((column, day))

        if day_columns:
            return row, day_columns

    return None, []


def _parse_sheet_month_year(worksheet):
    scan_rows = min(6, worksheet.max_row)
    scan_columns = min(8, worksheet.max_column)

    for row in range(1, scan_rows + 1):
        for column in range(1, scan_columns + 1):
            value = worksheet.cell(row, column).value
            if value is None:
                continue

            matched = SHEET_MONTH_PATTERN.search(str(value))
            if not matched:
                continue

            month = int(matched.group(1))
            year = int(matched.group(2))
            if 1 <= month <= 12:
                return year, month

    return None


def _purge_schedule_month(month_key):
    start_date, end_date = parse_month_key(month_key)

    schedule_ids = [
        row[0]
        for row in db.session.query(WorkSchedule.id)
        .filter(WorkSchedule.month_key == month_key)
        .all()
    ]

    deleted_overtime = 0
    if schedule_ids:
        deleted_overtime = OvertimeEntry.query.filter(
            OvertimeEntry.schedule_id.in_(schedule_ids)
        ).delete(synchronize_session=False)

    deleted_schedules = WorkSchedule.query.filter_by(month_key=month_key).delete(
        synchronize_session=False
    )
    deleted_details = AttendanceDetail.query.filter(
        AttendanceDetail.work_date >= start_date,
        AttendanceDetail.work_date <= end_date,
    ).delete(synchronize_session=False)

    return {
        "month_key": month_key,
        "deleted_schedules": int(deleted_schedules),
        "deleted_overtime": int(deleted_overtime),
        "deleted_details": int(deleted_details),
    }


def import_schedule_file(
    file_path,
    source_name,
    actor="system",
    target_month=None,
    replace_existing=True,
    blocked_month_keys=None,
):
    workbook = load_workbook(file_path, data_only=True)
    worksheet = _resolve_schedule_worksheet(workbook)

    shift_map = {row.code.upper(): row for row in ShiftTemplate.query.all()}
    if not shift_map:
        raise ValueError("Chua co bang ma ca lam")

    entries = []
    touched_months = set()
    invalid_shift_rows = []
    default_shift_applied = 0

    grid_header_row, day_columns = _detect_grid_schedule_layout(worksheet)

    if grid_header_row:
        month_year = _parse_sheet_month_year(worksheet)
        if target_month and month_year is None:
            target_start, _ = parse_month_key(target_month)
            month_year = (target_start.year, target_start.month)

        if month_year is None:
            raise ValueError(
                "Khong xac dinh duoc thang trong sheet Lich Lam. Vui long them tieu de co dang MM/YYYY"
            )

        year, month = month_year

        day_date_columns = []
        for column, day in day_columns:
            try:
                day_date_columns.append((column, date(year, month, day)))
            except ValueError:
                continue

        if not day_date_columns:
            raise ValueError("Sheet Lich Lam khong co cot ngay hop le cho thang duoc chon")

        employee_rows = []
        for row in range(grid_header_row + 1, worksheet.max_row + 1):
            employee_code = _coerce_employee_code(worksheet.cell(row, 2).value)
            employee_name_value = worksheet.cell(row, 3).value
            employee_name = (
                str(employee_name_value).strip() if employee_name_value is not None else ""
            )

            if not employee_code and not employee_name:
                continue
            if not employee_code:
                continue

            employee_rows.append((row, employee_code, employee_name))

        if not employee_rows:
            raise ValueError("Khong tim thay dong nhan vien trong sheet Lich Lam")

        employee_codes = sorted({item[1] for item in employee_rows})
        existing_employees = Employee.query.filter(Employee.employee_code.in_(employee_codes)).all()
        employee_map = {row.employee_code: row for row in existing_employees}

        for _, employee_code, employee_name in employee_rows:
            if employee_code in employee_map:
                continue

            employee = Employee(
                employee_code=employee_code,
                full_name=employee_name or f"Nhan vien {employee_code}",
                default_shift_code="X",
            )
            db.session.add(employee)
            db.session.flush()
            employee_map[employee_code] = employee
            log_action(
                "employees",
                employee_code,
                "INSERT",
                changed_by=actor,
                after_data=employee.to_dict(),
                notes="Tu dong tao nhan vien tu file lich lam",
            )

        for _, work_date in day_date_columns:
            row_month = month_key_for_date(work_date)
            if target_month and row_month != target_month:
                continue
            touched_months.add(row_month)

        for row, employee_code, _ in employee_rows:
            employee = employee_map[employee_code]

            for column, work_date in day_date_columns:
                row_month = month_key_for_date(work_date)
                if target_month and row_month != target_month:
                    continue

                cell_value = worksheet.cell(row, column).value
                shift_code = _normalize_shift_code(cell_value)
                if not shift_code:
                    if not _is_blank_cell(cell_value):
                        continue

                    shift_code = _normalize_shift_code(employee.default_shift_code)
                    if not shift_code:
                        continue

                    default_shift_applied += 1

                if shift_code not in shift_map:
                    invalid_label = "ma ca"
                    if _is_blank_cell(cell_value):
                        invalid_label = "ca mac dinh"
                    invalid_shift_rows.append(
                        f"Ngay {work_date} - NV {employee_code}: {invalid_label} '{shift_code}' khong ton tai"
                    )
                    continue

                entries.append((employee_code, work_date, shift_code))
    else:
        employee_columns = []
        for column in range(1, worksheet.max_column + 1):
            parsed = _parse_employee_header(worksheet.cell(1, column).value)
            if not parsed:
                continue
            employee_columns.append((column, parsed[0], parsed[1]))

        if not employee_columns:
            raise ValueError(
                "Khong tim thay du lieu lich hop le. Sheet can dung dinh dang ID <ma> - <ten> hoac Ma NV/Ho Ten"
            )

        employee_codes = [item[1] for item in employee_columns]
        existing_employees = Employee.query.filter(Employee.employee_code.in_(employee_codes)).all()
        employee_map = {row.employee_code: row for row in existing_employees}

        for _, employee_code, employee_name in employee_columns:
            if employee_code in employee_map:
                continue

            employee = Employee(
                employee_code=employee_code,
                full_name=employee_name,
                default_shift_code="X",
            )
            db.session.add(employee)
            db.session.flush()
            employee_map[employee_code] = employee
            log_action(
                "employees",
                employee_code,
                "INSERT",
                changed_by=actor,
                after_data=employee.to_dict(),
                notes="Tu dong tao nhan vien tu file lich lam",
            )

        for row in range(2, worksheet.max_row + 1):
            work_date = _coerce_date(worksheet.cell(row, 2).value)
            if not work_date:
                continue

            row_month = month_key_for_date(work_date)
            if target_month and row_month != target_month:
                continue

            touched_months.add(row_month)

            for column, employee_code, _ in employee_columns:
                cell_value = worksheet.cell(row, column).value
                shift_code = _normalize_shift_code(cell_value)
                if not shift_code:
                    if not _is_blank_cell(cell_value):
                        continue

                    employee = employee_map[employee_code]
                    shift_code = _normalize_shift_code(employee.default_shift_code)
                    if not shift_code:
                        continue

                    default_shift_applied += 1

                if shift_code not in shift_map:
                    invalid_label = "ma ca"
                    if _is_blank_cell(cell_value):
                        invalid_label = "ca mac dinh"
                    invalid_shift_rows.append(
                        f"Ngay {work_date} - NV {employee_code}: {invalid_label} '{shift_code}' khong ton tai"
                    )
                    continue

                entries.append((employee_code, work_date, shift_code))

    if invalid_shift_rows:
        preview = "; ".join(invalid_shift_rows[:15])
        raise ValueError(f"File lich co ma ca sai. Chi tiet: {preview}")

    touched_month_list = sorted(touched_months)
    if not touched_month_list:
        if target_month:
            raise ValueError(f"Khong tim thay du lieu lich cho thang {target_month}")
        raise ValueError("Khong tim thay du lieu lich hop le trong file")

    blocked_month_keys = {item for item in (blocked_month_keys or []) if item}
    locked_touched_months = sorted(set(touched_month_list).intersection(blocked_month_keys))
    if locked_touched_months:
        raise ValueError(
            "Khong the import lich vi cac thang da chot so: "
            + ", ".join(locked_touched_months)
        )

    replaced_months = []
    if replace_existing:
        for month_key in touched_month_list:
            replaced_months.append(_purge_schedule_month(month_key))

    created_count = 0
    updated_count = 0

    for employee_code, work_date, shift_code in entries:
        employee = employee_map[employee_code]
        shift = shift_map[shift_code]
        month_key = month_key_for_date(work_date)

        row = WorkSchedule.query.filter_by(employee_id=employee.id, work_date=work_date).first()
        if row:
            row.shift_id = shift.id
            row.month_key = month_key
            updated_count += 1
        else:
            row = WorkSchedule(
                employee_id=employee.id,
                work_date=work_date,
                month_key=month_key,
                shift_id=shift.id,
                absence_hours=0,
                notes=None,
            )
            db.session.add(row)
            created_count += 1

    batch_id = str(uuid.uuid4())
    log_action(
        "work_schedule_import",
        batch_id,
        "IMPORT",
        changed_by=actor,
        after_data={
            "source_file": source_name,
            "target_month": target_month,
            "months": touched_month_list,
            "replace_existing": bool(replace_existing),
            "replaced_months": replaced_months,
            "rows_imported": len(entries),
            "created": created_count,
            "updated": updated_count,
            "default_shift_applied": default_shift_applied,
        },
        notes="Import lich lam tu file xlsx",
    )

    db.session.commit()

    return {
        "batch_id": batch_id,
        "months": touched_month_list,
        "replace_existing": bool(replace_existing),
        "replaced_months": replaced_months,
        "rows_imported": len(entries),
        "created": created_count,
        "updated": updated_count,
        "default_shift_applied": default_shift_applied,
    }
