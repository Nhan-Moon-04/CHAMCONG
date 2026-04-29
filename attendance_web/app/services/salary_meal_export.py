import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import joinedload

from ..models import AttendanceDetail, Employee, ShiftTemplate
from .attendance import parse_month_key
from .nu_shift import is_nu_dynamic_shift_code


def _to_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _employee_code_sort_key(employee_code):
    raw_code = str(employee_code or "").replace("'", "").strip()
    if raw_code.isdigit():
        return (0, int(raw_code))
    return (1, raw_code.lower())


def _normalize_period(period):
    return 2 if str(period or "").strip() == "2" else 1


def _is_female(employee):
    """Check if employee is female (NU)"""
    if not employee or not employee.gender:
        return False
    gender = str(employee.gender or "").strip().upper()
    return "NU" in gender or "NỮ" in gender or gender == "F"


def _is_nu_night_row(detail_row):
    shift_code = str(getattr(detail_row, "shift_code", "") or "").strip().upper()
    if shift_code != "NU":
        return False

    shift_name = str(getattr(detail_row, "shift_name", "") or "").strip().lower()
    if "toi" in shift_name or "night" in shift_name:
        return True

    notes = str(getattr(detail_row, "notes", "") or "").strip().lower()
    return "toi" in notes or "night" in notes


def _get_meal_count_for_row(detail_row, shift_template):
    shift_code = str(getattr(detail_row, "shift_code", "") or "").strip().upper()

    if shift_code == "NU":
        return 1 if _is_nu_night_row(detail_row) else 2

    if shift_template:
        # Respect explicit zero meal_count on shift templates (do not coerce 0 -> 1)
        try:
            val = getattr(shift_template, "meal_count", None)
            if val is None:
                return 1
            return int(val)
        except (TypeError, ValueError):
            return 1

    return 1


def _get_meal_allowance_for_row(detail_row, shift_template):
    shift_code = str(getattr(detail_row, "shift_code", "") or "").strip().upper()

    if shift_code == "NU":
        return 35000.0

    if shift_template:
        return float(shift_template.meal_allowance or 0)

    return 0.0


def _apply_dash_zero_format(cell):
    cell.number_format = '#,##0;-#,##0;" - "'


def collect_salary_meal_overview_data(month_key, period, search_query=""):
    period = _normalize_period(period)
    search_query = (search_query or "").strip()

    start_date, end_date = parse_month_key(month_key)
    period_1_end = date(start_date.year, start_date.month, 15)
    period_2_start = date(start_date.year, start_date.month, 16)

    if period == 2:
        period_start = period_2_start
        period_end = end_date
        period_label = f"{period_start.strftime('%d/%m')} - {period_end.strftime('%d/%m')}"
        period_title = "Tien an dot 2"
    else:
        period_start = start_date
        period_end = period_1_end
        period_label = f"{period_start.strftime('%d/%m')} - {period_end.strftime('%d/%m')}"
        period_title = "Tien an dot 1"

    employees = Employee.query.filter(Employee.is_active.is_(True)).order_by(Employee.employee_code.asc()).all()
    meal_summary_map = {
        row.id: {
            "employee": row,
            "meal_count": 0,
            "meal_allowance": 0.0,
            "night_shift_count": 0,
        }
        for row in employees
    }

    detail_rows = (
        AttendanceDetail.query.options(joinedload(AttendanceDetail.employee))
        .filter(
            AttendanceDetail.month_key == month_key,
            AttendanceDetail.work_date >= period_start,
            AttendanceDetail.work_date <= period_end,
        )
        .order_by(AttendanceDetail.employee_id.asc(), AttendanceDetail.work_date.asc())
        .all()
    )

    # Build a map of shift code to shift template for night shift detection
    shift_templates = ShiftTemplate.query.all()
    shift_map = {st.code: st for st in shift_templates}

    for detail_row in detail_rows:
        if not detail_row.employee or not detail_row.shift_code:
            continue

        meal_summary = meal_summary_map.get(detail_row.employee_id)
        if not meal_summary:
            meal_summary = {
                "employee": detail_row.employee,
                "meal_count": 0,
                "meal_allowance": 0.0,
                "night_shift_count": 0,
            }
            meal_summary_map[detail_row.employee_id] = meal_summary

        # Skip leave/off codes based on status or shift template flags
        status = str(detail_row.status_code or "").upper()
        if status == "OFF":
            continue

        # Check shift template (may mark leave codes like OFF/O)
        shift_template = shift_map.get(detail_row.shift_code)
        if shift_template and getattr(shift_template, "is_leave_code", False):
            # explicit leave shift template: do not count meals
            continue

        # Count meals from the shift template, with NU morning/night using the special rule.
        # Also skip rows where the computed meal count is zero (e.g., OFF shifts with meal_count=0).
        if status not in {"P", "N"}:  # P=paid leave, N=unpaid leave
            shift_meal_count = _get_meal_count_for_row(detail_row, shift_template)
            if int(shift_meal_count or 0) == 0:
                continue
            meal_summary["meal_count"] += shift_meal_count
            meal_summary["meal_allowance"] = max(
                meal_summary["meal_allowance"],
                _get_meal_allowance_for_row(detail_row, shift_template),
            )
        
        # Count night shifts only for NU night rows.
        if _is_nu_night_row(detail_row):
            meal_summary["night_shift_count"] += 1

    meal_rows = []
    for employee_id, meal_summary in meal_summary_map.items():
        employee = meal_summary["employee"]
        
        # For female employees and NU shifts, keep the meal allowance at 35000.
        meal_allowance = meal_summary["meal_allowance"]
        if ("NU" in str(employee.gender or "").upper() or _is_female(employee)) and meal_allowance == 0:
            meal_allowance = 35000.0

        meal_total = float(meal_summary["meal_count"]) * float(meal_allowance)
        night_total = float(meal_summary["night_shift_count"]) * 100000.0
        
        meal_rows.append(
            {
                "employee_id": employee_id,
                "employee": employee,
                "meal_count": meal_summary["meal_count"],
                "meal_allowance": meal_allowance,
                "night_shift_count": meal_summary["night_shift_count"],
                # Backward-compatible fields for the existing meal page template.
                "worked_days": meal_summary["meal_count"],
                "paid_leave_days": 0.0,
                "unpaid_leave_days": 0.0,
                "meal_amount": meal_total + night_total,
            }
        )

    meal_rows.sort(
        key=lambda item: _employee_code_sort_key(item["employee"].employee_code)
    )

    if search_query:
        search_text = search_query.lower()

        def _match_meal_row(item):
            values = [
                item["employee"].employee_code,
                item["employee"].full_name,
                item["meal_count"],
                item["meal_allowance"],
                item["night_shift_count"],
            ]
            return any(
                search_text in str(value).lower()
                for value in values
                if value is not None
            )

        meal_rows = [item for item in meal_rows if _match_meal_row(item)]

    return {
        "month_key": month_key,
        "period": period,
        "period_title": period_title,
        "period_label": period_label,
        "period_start": period_start,
        "period_end": period_end,
        "search_query": search_query,
        "meal_rows": meal_rows,
    }


def build_salary_meal_export_excel(meal_data):
    """
    Build Excel file matching the template structure:
    STT, MSNV, HỌ VÀ TÊN, Số Bửa (meals), Tiền Cơm, Cộng Tiền cơm,
    Số Đêm, Bồi Dưỡng Ca Đêm, Cộng tiền bồi dưỡng Đêm,
    Số Bửa (aux), Tiền BD Phụ Xe, Cộng tiền bồi dưỡng,
    TIỀN ĐIỆN, TIỀN THỰC LÃNH, HỌ VÀ TÊN
    """
    workbook = Workbook()
    sheet = workbook.active
    
    period_start = meal_data.get("period_start")
    period_end = meal_data.get("period_end")
    period_label = f"{period_start.strftime('%d/%m/%Y')} - {period_end.strftime('%d/%m/%Y')}" if period_start else ""
    
    sheet.title = f"Tien an dot {meal_data['period']}"
    
    # Company info rows
    sheet.append(["CÔNG TY TNHH HIỆP LỢI"])
    sheet.append(["MST: 3701609885"])
    sheet.append([f"TIỀN CƠM VÀ TIỀN BỒI DƯỠNG TĂNG CA ĐÊM TỪ NGÀY {period_label}"])
    sheet.append([])
    
    # Header row
    headers = [
        "STT",
        "MSNV",
        "HỌ VÀ TÊN",
        "Số Bửa",
        "Tiền Cơm",
        "Cộng Tiền cơm",
        "Số Đêm",
        "Bồi Dưỡng Ca Đêm",
        "Cộng tiền bồi dưỡng Đêm",
        "Số Bửa",
        "Tiền BD Đi Phụ Xe",
        "Cộng tiền bồi dưỡng",
        "TIỀN ĐIỆN",
        "TIỀN THỰC LÃNH",
        "KÝ NHẬN",
    ]
    sheet.append(headers)
    header_row = 5
    
    # Format header
    header_font = Font(bold=True, size=10)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    for cell in sheet[header_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Add data rows
    for index, row in enumerate(meal_data["meal_rows"], start=1):
        employee = row["employee"]
        meal_count = row["meal_count"]
        meal_allowance = row["meal_allowance"]
        night_shift_count = row["night_shift_count"]
        
        # Row number
        row_num = header_row + index
        
        # STT
        sheet.cell(row=row_num, column=1).value = index
        _apply_dash_zero_format(sheet.cell(row=row_num, column=1))
        
        # MSNV (employee code)
        sheet.cell(row=row_num, column=2).value = employee.employee_code
        
        # HỌ VÀ TÊN (full name)
        sheet.cell(row=row_num, column=3).value = employee.full_name
        
        # Số Bửa (meal count)
        sheet.cell(row=row_num, column=4).value = meal_count
        _apply_dash_zero_format(sheet.cell(row=row_num, column=4))
        
        # Tiền Cơm (rice money)
        sheet.cell(row=row_num, column=5).value = meal_allowance
        _apply_dash_zero_format(sheet.cell(row=row_num, column=5))
        
        # Cộng Tiền cơm (total rice = col4 * col5)
        col_d = get_column_letter(4)
        col_e = get_column_letter(5)
        sheet.cell(row=row_num, column=6).value = f"={col_d}{row_num}*{col_e}{row_num}"
        _apply_dash_zero_format(sheet.cell(row=row_num, column=6))
        
        # Số Đêm (night shift count)
        sheet.cell(row=row_num, column=7).value = night_shift_count
        _apply_dash_zero_format(sheet.cell(row=row_num, column=7))
        
        # Bồi Dưỡng Ca Đêm (night allowance per night - fixed 100k)
        # Only populate when the employee actually worked night shifts in the period.
        if int(night_shift_count or 0) > 0:
            sheet.cell(row=row_num, column=8).value = 100000
        else:
            # write zero so the dash-format displays " - " for no allowance
            sheet.cell(row=row_num, column=8).value = 0
        _apply_dash_zero_format(sheet.cell(row=row_num, column=8))
        
        # Cộng tiền bồi dưỡng Đêm (total night = col7 * col8)
        col_g = get_column_letter(7)
        col_h = get_column_letter(8)
        sheet.cell(row=row_num, column=9).value = f"={col_g}{row_num}*{col_h}{row_num}"
        _apply_dash_zero_format(sheet.cell(row=row_num, column=9))
        
        # Số Bửa for auxiliary work stays blank in this export.
        sheet.cell(row=row_num, column=10).value = None
        
        # Tiền BD Đi Phụ Xe stays blank in this export.
        sheet.cell(row=row_num, column=11).value = None
        
        # Cộng tiền bồi dưỡng (total auxiliary = col10 * col11)
        col_j = get_column_letter(10)
        col_k = get_column_letter(11)
        sheet.cell(row=row_num, column=12).value = f"={col_j}{row_num}*{col_k}{row_num}"
        _apply_dash_zero_format(sheet.cell(row=row_num, column=12))
        
        # TIỀN ĐIỆN (electricity - blank)
        sheet.cell(row=row_num, column=13).value = None
        _apply_dash_zero_format(sheet.cell(row=row_num, column=13))
        
        # TIỀN THỰC LÃNH (actual received = col6 + col9 + col12 - col13)
        col_f = get_column_letter(6)
        col_i = get_column_letter(9)
        col_l = get_column_letter(12)
        col_m = get_column_letter(13)
        sheet.cell(row=row_num, column=14).value = f"={col_f}{row_num}+{col_i}{row_num}+{col_l}{row_num}-{col_m}{row_num}"
        _apply_dash_zero_format(sheet.cell(row=row_num, column=14))
        sheet.cell(row=row_num, column=14).font = Font(bold=True)
        
        # Ký nhận: để trống để người lao động tự ký.
        sheet.cell(row=row_num, column=15).value = None
    
    # Set column widths
    sheet.column_dimensions["A"].width = 5
    sheet.column_dimensions["B"].width = 10
    sheet.column_dimensions["C"].width = 20
    sheet.column_dimensions["D"].width = 10
    sheet.column_dimensions["E"].width = 12
    sheet.column_dimensions["F"].width = 14
    sheet.column_dimensions["G"].width = 8
    sheet.column_dimensions["H"].width = 12
    sheet.column_dimensions["I"].width = 14
    sheet.column_dimensions["J"].width = 10
    sheet.column_dimensions["K"].width = 14
    sheet.column_dimensions["L"].width = 14
    sheet.column_dimensions["M"].width = 10
    sheet.column_dimensions["N"].width = 14
    sheet.column_dimensions["O"].width = 20
    
    sheet.freeze_panes = "A6"
    
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    month_label = str(meal_data["month_key"]).replace("-", "")
    filename = f"tien_an_dot_{meal_data['period']}_{month_label}.xlsx"
    return output, filename
