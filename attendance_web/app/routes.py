import csv
import importlib
import io
import os
from datetime import date, datetime, timedelta
from pathlib import Path
from urllib.parse import urlparse

from flask import (
    Response,
    current_app,
    flash,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from sqlalchemy import String, cast, desc, func, or_
from sqlalchemy.orm import joinedload
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

from .database import db
from .models import (
    AdvancePayment,
    AppUser,
    AttendanceDaily,
    AttendanceDetail,
    AttendanceLog,
    AuditLog,
    Employee,
    Holiday,
    LeaveBalance,
    MonthlySalary,
    MonthlyWorkdayConfig,
    OvertimeEntry,
    ShiftTemplate,
    WorkSchedule,
)
from .services.attendance import (
    build_live_month_details,
    current_month_key,
    month_key_for_date,
    parse_month_key,
    rebuild_month_details,
)
from .services.audit import log_action
from .services.backup import run_pg_dump
from .services.importer import import_attendance_file
from .services.salary_importer import import_salary_file
from .services.schedule_importer import import_schedule_file


_holiday_lib = None
_holiday_lib_checked = False

DETAILS_HIGHLIGHT_TO_EXCEL_FILL = {
    "half-leave": "FFFFF1D6",
    "paid-leave": "FFE8F5E9",
    "unexcused": "FFFFB3B3",
    "missing-check": "FFFFF8CC",
}


def _safe_month_key(value):
    if not value:
        return current_month_key()
    try:
        datetime.strptime(value, "%Y-%m")
        return value
    except ValueError:
        return current_month_key()


def _parse_date(value):
    return datetime.strptime(value, "%Y-%m-%d").date()


def _parse_time(value):
    if not value:
        return None
    return datetime.strptime(value, "%H:%M").time()


def _to_float(value, default=0.0):
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _safe_advance_filter(value):
    normalized = (value or "all").strip().lower()
    if normalized in {"all", "has", "none"}:
        return normalized
    return "all"


def _employee_code_sort_key(employee_code):
    raw_code = str(employee_code or "").replace("'", "").strip()
    if raw_code.isdigit():
        return (0, int(raw_code))
    return (1, raw_code.lower())


def _safe_payment_method(value):
    normalized = (value or "cash").strip().lower()
    if normalized in {"cash", "salary_day"}:
        return normalized
    return "cash"


def _sanitize_next_path(value):
    if not value:
        return None

    parsed = urlparse(value)
    if parsed.scheme or parsed.netloc:
        return None

    if not value.startswith("/") or value.startswith("//"):
        return None

    return value


def _resolve_upload_relpath(relpath):
    if not relpath:
        return None

    try:
        upload_root = Path(current_app.config["UPLOAD_FOLDER"]).resolve()
        resolved_path = (upload_root / str(relpath)).resolve()
        resolved_path.relative_to(upload_root)
    except Exception:
        return None

    return resolved_path


def _iter_month_keys(start_date, end_date):
    current = date(start_date.year, start_date.month, 1)
    end_month = date(end_date.year, end_date.month, 1)

    while current <= end_month:
        yield current.strftime("%Y-%m")
        if current.month == 12:
            current = date(current.year + 1, 1, 1)
        else:
            current = date(current.year, current.month + 1, 1)


def _resolve_company_work_days(month_key):
    config = MonthlyWorkdayConfig.query.filter_by(month_key=month_key).first()
    config_value = _to_float(config.company_work_days if config else None, 0)
    if config_value > 0:
        return config_value, config

    legacy_salary = MonthlySalary.query.filter_by(month_key=month_key).order_by(MonthlySalary.id.asc()).first()
    legacy_value = _to_float(legacy_salary.salary_coefficient if legacy_salary else None, 0)
    if legacy_value >= 10:
        return legacy_value, config

    return 26.0, config


def _get_holiday_library():
    global _holiday_lib
    global _holiday_lib_checked

    if not _holiday_lib_checked:
        try:
            _holiday_lib = importlib.import_module("holidays")
        except ImportError:
            _holiday_lib = None
        _holiday_lib_checked = True

    return _holiday_lib


def _get_vietnam_holiday_map(start_date, end_date):
    holiday_map = {}
    years = list(range(start_date.year, end_date.year + 1))
    holiday_library = _get_holiday_library()

    if holiday_library:
        vn_calendar = holiday_library.country_holidays("VN", years=years)
        for holiday_date, holiday_name in vn_calendar.items():
            if start_date <= holiday_date <= end_date:
                holiday_map[holiday_date] = str(holiday_name)

    if not holiday_map:
        fixed_holidays = {
            (1, 1): "Tết Dương lịch",
            (4, 30): "Ngày giải phóng miền Nam",
            (5, 1): "Ngày Quốc tế Lao động",
            (9, 2): "Quốc khánh",
        }
        for year in years:
            for (month, day), holiday_name in fixed_holidays.items():
                holiday_date = date(year, month, day)
                if start_date <= holiday_date <= end_date:
                    holiday_map[holiday_date] = holiday_name

    return holiday_map


def _is_missing_check_event(detail_row):
    check_in = getattr(detail_row, "check_in", None)
    check_out = getattr(detail_row, "check_out", None)

    if check_in and check_out:
        return check_in == check_out
    return bool(check_in) != bool(check_out)


def _get_details_highlight_tag(detail_row):
    status_code = str(getattr(detail_row, "status_code", "") or "").upper()

    if _is_missing_check_event(detail_row):
        return "missing-check"
    if status_code in {"S", "C"}:
        return "half-leave"
    if status_code == "P":
        return "paid-leave"
    if status_code == "N":
        return "unexcused"
    return ""


def _collect_details_view_data(query_args, emit_flash=True):
    month_key = _safe_month_key(query_args.get("month"))
    employee_id_raw = (query_args.get("employee_id", "") or "").strip()
    search_query = (query_args.get("q", "") or "").strip()
    search_scope = (query_args.get("scope") or "current").strip().lower()
    if search_scope not in {"current", "all"}:
        search_scope = "current"

    selected_employee_id = None
    if employee_id_raw:
        try:
            selected_employee_id = int(employee_id_raw)
        except ValueError:
            if emit_flash:
                flash("Nhân viên không hợp lệ", "error")

    start_date_raw = (query_args.get("start_date", "") or "").strip()
    end_date_raw = (query_args.get("end_date", "") or "").strip()

    parsed_start_date = None
    parsed_end_date = None
    has_date_parse_error = False

    if start_date_raw:
        try:
            parsed_start_date = _parse_date(start_date_raw)
        except (TypeError, ValueError):
            if emit_flash:
                flash("Từ ngày không hợp lệ", "error")
            has_date_parse_error = True

    if end_date_raw:
        try:
            parsed_end_date = _parse_date(end_date_raw)
        except (TypeError, ValueError):
            if emit_flash:
                flash("Đến ngày không hợp lệ", "error")
            has_date_parse_error = True

    if has_date_parse_error:
        parsed_start_date = None
        parsed_end_date = None

    if parsed_start_date and not parsed_end_date:
        parsed_end_date = parsed_start_date
    if parsed_end_date and not parsed_start_date:
        parsed_start_date = parsed_end_date

    is_range_mode = bool(parsed_start_date and parsed_end_date)

    if is_range_mode and parsed_start_date > parsed_end_date:
        parsed_start_date, parsed_end_date = parsed_end_date, parsed_start_date

    def _sort_key(detail_row):
        raw_code = (detail_row.employee.employee_code or "").replace("'", "").strip()
        if raw_code.isdigit():
            code_key = (0, int(raw_code))
        else:
            code_key = (1, raw_code.lower())

        return code_key, detail_row.work_date, detail_row.employee.id

    if is_range_mode:
        rows = []
        for item_month in _iter_month_keys(parsed_start_date, parsed_end_date):
            month_rows = build_live_month_details(item_month, employee_id=selected_employee_id)
            rows.extend(
                row for row in month_rows if parsed_start_date <= row.work_date <= parsed_end_date
            )
        rows.sort(key=_sort_key)
        period_label = f"{parsed_start_date} đến {parsed_end_date}"
    else:
        if search_scope == "all":
            month_rows = (
                db.session.query(AttendanceDetail.month_key)
                .distinct()
                .order_by(AttendanceDetail.month_key.desc())
                .all()
            )
            month_keys = [row[0] for row in month_rows if row[0]]

            rows = []
            for item_month in month_keys:
                rows.extend(build_live_month_details(item_month, employee_id=selected_employee_id))

            rows.sort(key=_sort_key)
            period_label = "Toàn bộ dữ liệu"
        else:
            rows = build_live_month_details(month_key, employee_id=selected_employee_id)
            period_label = month_key

    if search_query:
        search_text = search_query.lower()

        def _detail_matches_search(detail_row):
            values = [
                getattr(detail_row.employee, "employee_code", ""),
                getattr(detail_row.employee, "full_name", ""),
                getattr(detail_row, "work_date", ""),
                getattr(detail_row, "shift_code", ""),
                getattr(detail_row, "shift_name", ""),
                getattr(detail_row, "check_in", ""),
                getattr(detail_row, "check_out", ""),
                getattr(detail_row, "actual_work_hours", ""),
                getattr(detail_row, "deviation_hours", ""),
                getattr(detail_row, "overtime_hours", ""),
                getattr(detail_row, "total_span_hours", ""),
                getattr(detail_row, "status_code", ""),
                getattr(detail_row, "paid_hours", ""),
                getattr(detail_row, "daily_wage", ""),
                getattr(detail_row, "notes", ""),
                getattr(detail_row, "meal_allowance_daily", ""),
            ]

            return any(
                search_text in str(value).lower()
                for value in values
                if value is not None
            )

        rows = [row for row in rows if _detail_matches_search(row)]

    for row in rows:
        setattr(row, "highlight_tag", _get_details_highlight_tag(row))

    query_params = {"month": month_key}
    if selected_employee_id is not None:
        query_params["employee_id"] = selected_employee_id
    if parsed_start_date:
        query_params["start_date"] = parsed_start_date.isoformat()
    if parsed_end_date:
        query_params["end_date"] = parsed_end_date.isoformat()
    if search_scope == "all":
        query_params["scope"] = "all"
    if search_query:
        query_params["q"] = search_query

    return {
        "month_key": month_key,
        "rows": rows,
        "selected_employee_id": selected_employee_id,
        "parsed_start_date": parsed_start_date,
        "parsed_end_date": parsed_end_date,
        "is_range_mode": is_range_mode,
        "period_label": period_label,
        "search_query": search_query,
        "search_scope": search_scope,
        "query_params": query_params,
    }


def _normalize_username(value):
    return (value or "").strip().lower()


def _user_audit_payload(user):
    return {
        "id": user.id,
        "username": user.username,
        "full_name": user.full_name,
        "is_admin": bool(user.is_admin),
        "is_active": bool(user.is_active),
    }


def _active_admin_count():
    return AppUser.query.filter_by(is_admin=True, is_active=True).count()


def register_routes(app):
    def _set_auth_session(user):
        session["is_authenticated"] = True
        session["user_id"] = user.id
        session["username"] = user.username
        session["display_name"] = user.full_name or user.username
        session["is_admin"] = bool(user.is_admin)

    def _require_admin():
        if session.get("is_admin"):
            return None

        flash("Bạn không có quyền truy cập chức năng này", "error")
        return redirect(url_for("dashboard"))

    @app.context_processor
    def inject_auth_state():
        return {
            "is_authenticated": bool(session.get("is_authenticated")),
            "current_user": session.get("display_name") or session.get("username", ""),
            "current_user_username": session.get("username", ""),
            "current_user_is_admin": bool(session.get("is_admin")),
            "current_user_id": session.get("user_id"),
        }

    @app.before_request
    def require_login():
        endpoint = request.endpoint or ""
        if endpoint == "login" or endpoint == "static" or endpoint.startswith("static."):
            return None

        if not session.get("is_authenticated"):
            next_path = request.full_path if request.query_string else request.path
            return redirect(url_for("login", next=next_path))

        current_user = None
        user_id = session.get("user_id")
        username = _normalize_username(session.get("username", ""))

        if user_id is not None:
            current_user = AppUser.query.filter_by(id=user_id).first()

        if current_user is None and username:
            current_user = AppUser.query.filter(func.lower(AppUser.username) == username).first()

        if current_user is None or not current_user.is_active:
            session.clear()
            flash("Tài khoản không hợp lệ hoặc đã bị khóa", "error")
            next_path = request.full_path if request.query_string else request.path
            return redirect(url_for("login", next=next_path))

        _set_auth_session(current_user)

        return None

    @app.route("/login", methods=["GET", "POST"])
    def login():
        if session.get("is_authenticated"):
            return redirect(url_for("dashboard"))

        query_next = _sanitize_next_path((request.args.get("next") or "").strip())

        if request.method == "POST":
            username = _normalize_username(request.form.get("username"))
            password = request.form.get("password") or ""

            user = None
            if username:
                user = AppUser.query.filter(func.lower(AppUser.username) == username).first()

            if user and user.is_active and check_password_hash(user.password_hash, password):
                session.clear()
                _set_auth_session(user)
                session.permanent = True
                flash("Đăng nhập thành công", "success")

                form_next = _sanitize_next_path((request.form.get("next") or "").strip())
                target = form_next or query_next or url_for("dashboard")
                return redirect(target)

            if user and not user.is_active:
                flash("Tài khoản đã bị khóa", "error")
            else:
                flash("Sai tài khoản hoặc mật khẩu", "error")

        return render_template("login.html", title="Đăng nhập", next_url=query_next or "")

    @app.route("/logout", methods=["POST"])
    def logout():
        session.clear()
        flash("Đã đăng xuất", "success")
        return redirect(url_for("login"))

    @app.route("/users", methods=["GET"])
    def users():
        blocked = _require_admin()
        if blocked:
            return blocked

        rows = AppUser.query.order_by(AppUser.username.asc()).all()
        active_users = [row for row in rows if row.is_active]
        admin_users = [row for row in rows if row.is_admin]

        return render_template(
            "users.html",
            title="Quản lý user",
            users=rows,
            total_users=len(rows),
            active_users=len(active_users),
            admin_users=len(admin_users),
            current_user_id=session.get("user_id"),
        )

    @app.route("/users/new", methods=["GET", "POST"])
    def create_user():
        blocked = _require_admin()
        if blocked:
            return blocked

        form_values = {
            "username": "",
            "full_name": "",
            "is_admin": False,
            "is_active": True,
        }

        if request.method == "POST":
            actor = session.get("username") or "admin"
            username = _normalize_username(request.form.get("username"))
            full_name = (request.form.get("full_name") or "").strip() or None
            password = request.form.get("password") or ""
            is_admin = request.form.get("is_admin") == "on"
            is_active = request.form.get("is_active") == "on"

            form_values.update(
                {
                    "username": username,
                    "full_name": full_name or "",
                    "is_admin": is_admin,
                    "is_active": is_active,
                }
            )

            if not username:
                flash("Cần nhập tên đăng nhập", "error")
                return render_template(
                    "user_form.html",
                    title="Thêm user",
                    form_title="Thêm user mới",
                    form_subtitle="Tạo tài khoản đăng nhập mới cho hệ thống.",
                    submit_label="Tạo user",
                    back_url=url_for("users"),
                    form_values=form_values,
                    is_edit=False,
                )

            if len(password) < 4:
                flash("Mật khẩu tối thiểu 4 ký tự", "error")
                return render_template(
                    "user_form.html",
                    title="Thêm user",
                    form_title="Thêm user mới",
                    form_subtitle="Tạo tài khoản đăng nhập mới cho hệ thống.",
                    submit_label="Tạo user",
                    back_url=url_for("users"),
                    form_values=form_values,
                    is_edit=False,
                )

            existing = AppUser.query.filter(func.lower(AppUser.username) == username).first()
            if existing:
                flash("Tên đăng nhập đã tồn tại", "error")
                return render_template(
                    "user_form.html",
                    title="Thêm user",
                    form_title="Thêm user mới",
                    form_subtitle="Tạo tài khoản đăng nhập mới cho hệ thống.",
                    submit_label="Tạo user",
                    back_url=url_for("users"),
                    form_values=form_values,
                    is_edit=False,
                )

            user = AppUser(
                username=username,
                full_name=full_name,
                password_hash=generate_password_hash(password),
                is_admin=is_admin,
                is_active=is_active,
            )
            db.session.add(user)
            db.session.flush()

            log_action(
                "app_users",
                str(user.id),
                "INSERT",
                changed_by=actor,
                after_data=_user_audit_payload(user),
            )
            db.session.commit()
            flash("Đã thêm user", "success")
            return redirect(url_for("users"))

        return render_template(
            "user_form.html",
            title="Thêm user",
            form_title="Thêm user mới",
            form_subtitle="Tạo tài khoản đăng nhập mới cho hệ thống.",
            submit_label="Tạo user",
            back_url=url_for("users"),
            form_values=form_values,
            is_edit=False,
        )

    @app.route("/users/<int:user_id>/edit", methods=["GET", "POST"])
    def edit_user(user_id):
        blocked = _require_admin()
        if blocked:
            return blocked

        user = AppUser.query.get_or_404(user_id)
        form_values = {
            "username": user.username,
            "full_name": user.full_name or "",
            "is_admin": bool(user.is_admin),
            "is_active": bool(user.is_active),
        }

        if request.method == "POST":
            actor = session.get("username") or "admin"
            username = _normalize_username(request.form.get("username"))
            full_name = (request.form.get("full_name") or "").strip() or None
            password = request.form.get("password") or ""
            is_admin = request.form.get("is_admin") == "on"
            is_active = request.form.get("is_active") == "on"

            form_values.update(
                {
                    "username": username,
                    "full_name": full_name or "",
                    "is_admin": is_admin,
                    "is_active": is_active,
                }
            )

            if not username:
                flash("Cần nhập tên đăng nhập", "error")
                return render_template(
                    "user_form.html",
                    title="Sửa user",
                    form_title="Cập nhật user",
                    form_subtitle="Cập nhật quyền và thông tin tài khoản.",
                    submit_label="Lưu cập nhật",
                    back_url=url_for("users"),
                    form_values=form_values,
                    is_edit=True,
                    user=user,
                )

            if password and len(password) < 4:
                flash("Mật khẩu mới tối thiểu 4 ký tự", "error")
                return render_template(
                    "user_form.html",
                    title="Sửa user",
                    form_title="Cập nhật user",
                    form_subtitle="Cập nhật quyền và thông tin tài khoản.",
                    submit_label="Lưu cập nhật",
                    back_url=url_for("users"),
                    form_values=form_values,
                    is_edit=True,
                    user=user,
                )

            duplicate = (
                AppUser.query.filter(func.lower(AppUser.username) == username, AppUser.id != user.id)
                .first()
            )
            if duplicate:
                flash("Tên đăng nhập đã tồn tại", "error")
                return render_template(
                    "user_form.html",
                    title="Sửa user",
                    form_title="Cập nhật user",
                    form_subtitle="Cập nhật quyền và thông tin tài khoản.",
                    submit_label="Lưu cập nhật",
                    back_url=url_for("users"),
                    form_values=form_values,
                    is_edit=True,
                    user=user,
                )

            current_user_id = session.get("user_id")
            is_last_active_admin = user.is_admin and user.is_active and _active_admin_count() <= 1

            if is_last_active_admin and (not is_admin or not is_active):
                flash("Không thể hạ quyền user admin cuối cùng", "error")
                return render_template(
                    "user_form.html",
                    title="Sửa user",
                    form_title="Cập nhật user",
                    form_subtitle="Cập nhật quyền và thông tin tài khoản.",
                    submit_label="Lưu cập nhật",
                    back_url=url_for("users"),
                    form_values=form_values,
                    is_edit=True,
                    user=user,
                )

            if user.id == current_user_id and (not is_admin or not is_active):
                flash("Không thể tự khóa hoặc bỏ quyền admin của tài khoản đang đăng nhập", "error")
                return render_template(
                    "user_form.html",
                    title="Sửa user",
                    form_title="Cập nhật user",
                    form_subtitle="Cập nhật quyền và thông tin tài khoản.",
                    submit_label="Lưu cập nhật",
                    back_url=url_for("users"),
                    form_values=form_values,
                    is_edit=True,
                    user=user,
                )

            before_data = _user_audit_payload(user)

            user.username = username
            user.full_name = full_name
            user.is_admin = is_admin
            user.is_active = is_active
            if password:
                user.password_hash = generate_password_hash(password)

            db.session.flush()
            log_action(
                "app_users",
                str(user.id),
                "UPDATE",
                changed_by=actor,
                before_data=before_data,
                after_data=_user_audit_payload(user),
            )
            db.session.commit()

            if user.id == current_user_id:
                _set_auth_session(user)

            flash("Đã cập nhật user", "success")
            return redirect(url_for("users"))

        return render_template(
            "user_form.html",
            title="Sửa user",
            form_title="Cập nhật user",
            form_subtitle="Cập nhật quyền và thông tin tài khoản.",
            submit_label="Lưu cập nhật",
            back_url=url_for("users"),
            form_values=form_values,
            is_edit=True,
            user=user,
        )

    @app.route("/users/<int:user_id>/delete", methods=["POST"])
    def delete_user(user_id):
        blocked = _require_admin()
        if blocked:
            return blocked

        user = AppUser.query.get_or_404(user_id)
        current_user_id = session.get("user_id")
        if user.id == current_user_id:
            flash("Không thể xóa tài khoản đang đăng nhập", "error")
            return redirect(url_for("users"))

        if user.is_admin and user.is_active and _active_admin_count() <= 1:
            flash("Không thể xóa user admin cuối cùng", "error")
            return redirect(url_for("users"))

        actor = session.get("username") or "admin"
        before_data = _user_audit_payload(user)
        db.session.delete(user)

        log_action(
            "app_users",
            str(user.id),
            "DELETE",
            changed_by=actor,
            before_data=before_data,
        )
        db.session.commit()
        flash("Đã xóa user", "success")
        return redirect(url_for("users"))

    @app.route("/")
    def dashboard():
        month_key = _safe_month_key(request.args.get("month"))
        warning_query = (request.args.get("q") or "").strip()

        # Keep dashboard numbers in sync with source tables (schedule, shifts, attendance, salary).
        rebuild_month_details(month_key, actor="system-auto-sync", write_audit=False)

        total_employees = Employee.query.filter_by(is_active=True).count()
        detail_query = AttendanceDetail.query.filter_by(month_key=month_key)

        total_rows = detail_query.count()
        total_paid_hours = (
            db.session.query(func.coalesce(func.sum(AttendanceDetail.paid_hours), 0))
            .filter(AttendanceDetail.month_key == month_key)
            .scalar()
        )
        total_wage = (
            db.session.query(func.coalesce(func.sum(AttendanceDetail.daily_wage), 0))
            .filter(AttendanceDetail.month_key == month_key)
            .scalar()
        )
        absent_days = detail_query.filter(AttendanceDetail.status_code == "N").count()

        status_rows = (
            db.session.query(AttendanceDetail.status_code, func.count(AttendanceDetail.id))
            .filter(AttendanceDetail.month_key == month_key)
            .group_by(AttendanceDetail.status_code)
            .order_by(AttendanceDetail.status_code.asc())
            .all()
        )

        overtime_rows = (
            db.session.query(
                Employee.full_name,
                func.coalesce(func.sum(AttendanceDetail.overtime_hours), 0).label("ot_hours"),
            )
            .join(AttendanceDetail, AttendanceDetail.employee_id == Employee.id)
            .filter(AttendanceDetail.month_key == month_key)
            .group_by(Employee.full_name)
            .order_by(desc("ot_hours"))
            .limit(10)
            .all()
        )

        warning_rows_query = (
            db.session.query(AttendanceDetail, Employee)
            .join(Employee, AttendanceDetail.employee_id == Employee.id)
            .filter(AttendanceDetail.month_key == month_key)
            .filter(
                or_(
                    AttendanceDetail.status_code == "N",
                    AttendanceDetail.deviation_hours < 0,
                )
            )
        )

        if warning_query:
            warning_like = f"%{warning_query}%"
            warning_rows_query = warning_rows_query.filter(
                or_(
                    Employee.employee_code.ilike(warning_like),
                    Employee.full_name.ilike(warning_like),
                    AttendanceDetail.status_code.ilike(warning_like),
                    AttendanceDetail.notes.ilike(warning_like),
                    cast(AttendanceDetail.work_date, String).ilike(warning_like),
                    cast(AttendanceDetail.deviation_hours, String).ilike(warning_like),
                )
            )

        warning_total = warning_rows_query.count()
        warning_rows = (
            warning_rows_query
            .order_by(AttendanceDetail.work_date.desc(), AttendanceDetail.id.desc())
            .all()
        )

        attendance_rate = (
            ((total_rows - absent_days) / total_rows * 100.0) if total_rows > 0 else 0.0
        )
        average_paid_hours = (float(total_paid_hours or 0) / total_rows) if total_rows > 0 else 0.0
        average_daily_wage = (float(total_wage or 0) / total_rows) if total_rows > 0 else 0.0

        return render_template(
            "dashboard.html",
            title="Bảng điều khiển",
            month_key=month_key,
            total_employees=total_employees,
            total_rows=total_rows,
            total_paid_hours=float(total_paid_hours or 0),
            total_wage=float(total_wage or 0),
            absent_days=absent_days,
            warning_total=int(warning_total or 0),
            attendance_rate=attendance_rate,
            average_paid_hours=average_paid_hours,
            average_daily_wage=average_daily_wage,
            status_labels=[row[0] for row in status_rows],
            status_values=[int(row[1]) for row in status_rows],
            overtime_labels=[row[0] for row in overtime_rows],
            overtime_values=[float(row[1]) for row in overtime_rows],
            warning_rows=warning_rows,
            warning_query=warning_query,
        )

    @app.route("/employees", methods=["GET"])
    def employees():
        search_query = (request.args.get("q") or "").strip()
        search_scope = (request.args.get("scope") or "current").strip().lower()
        if search_scope not in {"current", "all"}:
            search_scope = "current"

        rows_query = Employee.query
        if search_scope == "current":
            rows_query = rows_query.filter(Employee.is_active.is_(True))

        if search_query:
            search_like = f"%{search_query}%"
            search_filters = [
                Employee.employee_code.ilike(search_like),
                Employee.full_name.ilike(search_like),
                Employee.gender.ilike(search_like),
                Employee.hometown.ilike(search_like),
                Employee.default_shift_code.ilike(search_like),
                cast(Employee.birth_year, String).ilike(search_like),
            ]

            query_lower = search_query.lower()
            if any(token in query_lower for token in ["hoat", "hoạt", "active", "current", "hiện tại"]):
                search_filters.append(Employee.is_active.is_(True))
            if any(token in query_lower for token in ["tam", "tạm", "ngung", "ngưng", "inactive"]):
                search_filters.append(Employee.is_active.is_(False))

            rows_query = rows_query.filter(or_(*search_filters))

        rows = rows_query.order_by(Employee.created_at.desc(), Employee.id.desc()).all()
        active_rows = [row for row in rows if row.is_active]
        return render_template(
            "employees.html",
            title="Nhân viên",
            employees=rows,
            total_employees=len(rows),
            active_employees=len(active_rows),
            search_query=search_query,
            search_scope=search_scope,
        )

    @app.route("/employees/new", methods=["GET", "POST"])
    def create_employee():
        shift_codes = [row.code for row in ShiftTemplate.query.order_by(ShiftTemplate.code.asc()).all()]

        form_values = {
            "employee_code": "",
            "full_name": "",
            "gender": "",
            "hometown": "",
            "birth_year": "",
            "default_shift_code": "X",
            "is_active": True,
            "changed_by": session.get("username", "admin"),
        }

        if request.method == "POST":
            actor = (request.form.get("changed_by") or session.get("username") or "admin").strip() or "admin"
            employee_code = request.form.get("employee_code", "").strip()
            full_name = request.form.get("full_name", "").strip()
            gender = request.form.get("gender", "").strip() or None
            hometown = request.form.get("hometown", "").strip() or None
            birth_year_raw = request.form.get("birth_year", "").strip()
            default_shift_code = request.form.get("default_shift_code", "X").strip().upper()
            is_active = request.form.get("is_active") == "on"

            form_values.update(
                {
                    "employee_code": employee_code,
                    "full_name": full_name,
                    "gender": gender or "",
                    "hometown": hometown or "",
                    "birth_year": birth_year_raw,
                    "default_shift_code": default_shift_code,
                    "is_active": is_active,
                    "changed_by": actor,
                }
            )

            if not employee_code or not full_name:
                flash("Cần nhập Mã NV và Họ tên", "error")
                return render_template(
                    "employee_form.html",
                    title="Thêm nhân viên",
                    form_title="Thêm nhân viên mới",
                    form_subtitle="Tạo hồ sơ nhân viên và gán ca mặc định.",
                    submit_label="Tạo nhân viên",
                    form_values=form_values,
                    shift_codes=shift_codes,
                    is_edit=False,
                )

            if default_shift_code not in shift_codes:
                flash("Mã ca mặc định không hợp lệ", "error")
                return render_template(
                    "employee_form.html",
                    title="Thêm nhân viên",
                    form_title="Thêm nhân viên mới",
                    form_subtitle="Tạo hồ sơ nhân viên và gán ca mặc định.",
                    submit_label="Tạo nhân viên",
                    form_values=form_values,
                    shift_codes=shift_codes,
                    is_edit=False,
                )

            existing = Employee.query.filter_by(employee_code=employee_code).first()
            if existing:
                flash("Mã nhân viên đã tồn tại", "error")
                return render_template(
                    "employee_form.html",
                    title="Thêm nhân viên",
                    form_title="Thêm nhân viên mới",
                    form_subtitle="Tạo hồ sơ nhân viên và gán ca mặc định.",
                    submit_label="Tạo nhân viên",
                    form_values=form_values,
                    shift_codes=shift_codes,
                    is_edit=False,
                )

            birth_year = None
            if birth_year_raw:
                if not birth_year_raw.isdigit():
                    flash("Năm sinh không hợp lệ", "error")
                    return render_template(
                        "employee_form.html",
                        title="Thêm nhân viên",
                        form_title="Thêm nhân viên mới",
                        form_subtitle="Tạo hồ sơ nhân viên và gán ca mặc định.",
                        submit_label="Tạo nhân viên",
                        form_values=form_values,
                        shift_codes=shift_codes,
                        is_edit=False,
                    )
                birth_year = int(birth_year_raw)

            employee = Employee(
                employee_code=employee_code,
                full_name=full_name,
                gender=gender,
                hometown=hometown,
                birth_year=birth_year,
                default_shift_code=default_shift_code,
                is_active=is_active,
            )
            db.session.add(employee)
            db.session.flush()

            log_action(
                "employees",
                employee.id,
                "INSERT",
                changed_by=actor,
                after_data=employee.to_dict(),
            )
            db.session.commit()
            flash("Đã thêm nhân viên", "success")
            return redirect(url_for("employees"))

        return render_template(
            "employee_form.html",
            title="Thêm nhân viên",
            form_title="Thêm nhân viên mới",
            form_subtitle="Tạo hồ sơ nhân viên và gán ca mặc định.",
            submit_label="Tạo nhân viên",
            form_values=form_values,
            shift_codes=shift_codes,
            is_edit=False,
        )

    @app.route("/employees/<int:employee_id>/edit", methods=["GET", "POST"])
    def edit_employee(employee_id):
        employee = Employee.query.get_or_404(employee_id)
        shift_codes = [row.code for row in ShiftTemplate.query.order_by(ShiftTemplate.code.asc()).all()]

        form_values = {
            "employee_code": employee.employee_code,
            "full_name": employee.full_name,
            "gender": employee.gender or "",
            "hometown": employee.hometown or "",
            "birth_year": employee.birth_year or "",
            "default_shift_code": employee.default_shift_code,
            "is_active": bool(employee.is_active),
            "changed_by": session.get("username", "admin"),
        }

        if request.method == "POST":
            actor = (request.form.get("changed_by") or session.get("username") or "admin").strip() or "admin"
            employee_code = request.form.get("employee_code", "").strip()
            full_name = request.form.get("full_name", "").strip()
            gender = request.form.get("gender", "").strip() or None
            hometown = request.form.get("hometown", "").strip() or None
            birth_year_raw = request.form.get("birth_year", "").strip()
            default_shift_code = request.form.get("default_shift_code", "X").strip().upper()
            is_active = request.form.get("is_active") == "on"

            form_values.update(
                {
                    "employee_code": employee_code,
                    "full_name": full_name,
                    "gender": gender or "",
                    "hometown": hometown or "",
                    "birth_year": birth_year_raw,
                    "default_shift_code": default_shift_code,
                    "is_active": is_active,
                    "changed_by": actor,
                }
            )

            if not employee_code or not full_name:
                flash("Cần nhập Mã NV và Họ tên", "error")
                return render_template(
                    "employee_form.html",
                    title="Sửa nhân viên",
                    form_title="Cập nhật thông tin nhân viên",
                    form_subtitle="Có thể sửa thông tin ca mặc định và trạng thái hoạt động.",
                    submit_label="Lưu cập nhật",
                    form_values=form_values,
                    shift_codes=shift_codes,
                    is_edit=True,
                    employee=employee,
                )

            if default_shift_code not in shift_codes:
                flash("Mã ca mặc định không hợp lệ", "error")
                return render_template(
                    "employee_form.html",
                    title="Sửa nhân viên",
                    form_title="Cập nhật thông tin nhân viên",
                    form_subtitle="Có thể sửa thông tin ca mặc định và trạng thái hoạt động.",
                    submit_label="Lưu cập nhật",
                    form_values=form_values,
                    shift_codes=shift_codes,
                    is_edit=True,
                    employee=employee,
                )

            duplicated = (
                Employee.query.filter(
                    Employee.employee_code == employee_code,
                    Employee.id != employee.id,
                ).first()
            )
            if duplicated:
                flash("Mã nhân viên đã tồn tại", "error")
                return render_template(
                    "employee_form.html",
                    title="Sửa nhân viên",
                    form_title="Cập nhật thông tin nhân viên",
                    form_subtitle="Có thể sửa thông tin ca mặc định và trạng thái hoạt động.",
                    submit_label="Lưu cập nhật",
                    form_values=form_values,
                    shift_codes=shift_codes,
                    is_edit=True,
                    employee=employee,
                )

            birth_year = None
            if birth_year_raw:
                if not birth_year_raw.isdigit():
                    flash("Năm sinh không hợp lệ", "error")
                    return render_template(
                        "employee_form.html",
                        title="Sửa nhân viên",
                        form_title="Cập nhật thông tin nhân viên",
                        form_subtitle="Có thể sửa thông tin ca mặc định và trạng thái hoạt động.",
                        submit_label="Lưu cập nhật",
                        form_values=form_values,
                        shift_codes=shift_codes,
                        is_edit=True,
                        employee=employee,
                    )
                birth_year = int(birth_year_raw)

            before = employee.to_dict()
            employee.employee_code = employee_code
            employee.full_name = full_name
            employee.gender = gender
            employee.hometown = hometown
            employee.birth_year = birth_year
            employee.default_shift_code = default_shift_code
            employee.is_active = is_active

            log_action(
                "employees",
                employee.id,
                "UPDATE",
                changed_by=actor,
                before_data=before,
                after_data=employee.to_dict(),
            )
            db.session.commit()
            flash("Đã cập nhật nhân viên", "success")
            return redirect(url_for("employees"))

        return render_template(
            "employee_form.html",
            title="Sửa nhân viên",
            form_title="Cập nhật thông tin nhân viên",
            form_subtitle="Có thể sửa thông tin ca mặc định và trạng thái hoạt động.",
            submit_label="Lưu cập nhật",
            form_values=form_values,
            shift_codes=shift_codes,
            is_edit=True,
            employee=employee,
        )

    @app.route("/employees/<int:employee_id>")
    def employee_detail(employee_id):
        employee = Employee.query.get_or_404(employee_id)
        month_key = _safe_month_key(request.args.get("month"))

        details = build_live_month_details(month_key, employee_id=employee_id)
        salaries = (
            MonthlySalary.query.filter_by(employee_id=employee_id)
            .order_by(MonthlySalary.month_key.desc())
            .all()
        )
        balances = (
            LeaveBalance.query.filter_by(employee_id=employee_id)
            .order_by(LeaveBalance.year.desc())
            .all()
        )

        month_keys = [row.month_key for row in salaries]
        config_rows = (
            MonthlyWorkdayConfig.query.filter(MonthlyWorkdayConfig.month_key.in_(month_keys)).all()
            if month_keys
            else []
        )
        config_map = {row.month_key: _to_float(row.company_work_days, 0) for row in config_rows}

        salary_history_rows = []
        for row in salaries:
            company_work_days = config_map.get(row.month_key, 0)
            if company_work_days <= 0:
                legacy_value = _to_float(row.salary_coefficient, 0)
                company_work_days = legacy_value if legacy_value >= 10 else 26.0

            monthly_wage = _to_float(row.base_daily_wage)
            daily_rate = (monthly_wage / company_work_days) if company_work_days > 0 else 0.0
            salary_history_rows.append(
                {
                    "salary": row,
                    "monthly_wage": round(monthly_wage, 2),
                    "company_work_days": round(company_work_days, 2),
                    "daily_rate": round(daily_rate, 2),
                }
            )

        summary_paid_hours = sum(float(row.paid_hours or 0) for row in details)
        summary_wage = sum(float(row.daily_wage or 0) for row in details)

        return render_template(
            "employee_detail.html",
            employee=employee,
            month_key=month_key,
            details=details,
            salary_history_rows=salary_history_rows,
            balances=balances,
            summary_paid_hours=summary_paid_hours,
            summary_wage=summary_wage,
        )

    @app.route("/shifts", methods=["GET", "POST"])
    def shifts():
        if request.method == "POST":
            actor = request.form.get("changed_by", "admin").strip() or "admin"
            code = request.form.get("code", "").strip().upper()
            name = request.form.get("name", "").strip()
            payload = {
                "name": name,
                "start_time": _parse_time(request.form.get("start_time", "")),
                "end_time": _parse_time(request.form.get("end_time", "")),
                "break_minutes": int(request.form.get("break_minutes", 0) or 0),
                "standard_hours": _to_float(request.form.get("standard_hours"), 0),
                "default_overtime_hours": _to_float(
                    request.form.get("default_overtime_hours"), 0
                ),
                "meal_allowance": _to_float(request.form.get("meal_allowance"), 0),
                "is_leave_code": request.form.get("is_leave_code") == "on",
                "is_paid_leave": request.form.get("is_paid_leave") == "on",
                "notes": request.form.get("notes", "").strip() or None,
            }

            if not code or not name:
                flash("Cần nhập mã ca và tên ca", "error")
                return redirect(url_for("shifts"))

            existing = ShiftTemplate.query.filter_by(code=code).first()
            if existing:
                flash("Mã ca đã tồn tại, vui lòng bấm Sửa trên dòng cần cập nhật", "error")
                return redirect(url_for("edit_shift", shift_id=existing.id))

            shift = ShiftTemplate(code=code, **payload)
            db.session.add(shift)
            db.session.flush()
            log_action(
                "shift_templates",
                shift.id,
                "INSERT",
                changed_by=actor,
                after_data=shift.to_dict(),
            )

            db.session.commit()
            flash("Đã tạo ca mới", "success")
            return redirect(url_for("shifts"))

        rows = ShiftTemplate.query.order_by(ShiftTemplate.code.asc()).all()
        return render_template("shifts.html", shifts=rows)

    @app.route("/shifts/<int:shift_id>/edit", methods=["GET", "POST"])
    def edit_shift(shift_id):
        shift = ShiftTemplate.query.get_or_404(shift_id)

        if request.method == "POST":
            actor = request.form.get("changed_by", "admin").strip() or "admin"
            name = request.form.get("name", "").strip()

            if not name:
                flash("Cần nhập tên ca", "error")
                return redirect(url_for("edit_shift", shift_id=shift.id))

            payload = {
                "name": name,
                "start_time": _parse_time(request.form.get("start_time", "")),
                "end_time": _parse_time(request.form.get("end_time", "")),
                "break_minutes": int(request.form.get("break_minutes", 0) or 0),
                "standard_hours": _to_float(request.form.get("standard_hours"), 0),
                "default_overtime_hours": _to_float(
                    request.form.get("default_overtime_hours"), 0
                ),
                "meal_allowance": _to_float(request.form.get("meal_allowance"), 0),
                "is_leave_code": request.form.get("is_leave_code") == "on",
                "is_paid_leave": request.form.get("is_paid_leave") == "on",
                "notes": request.form.get("notes", "").strip() or None,
            }

            before = shift.to_dict()
            for key, value in payload.items():
                setattr(shift, key, value)

            log_action(
                "shift_templates",
                shift.id,
                "UPDATE",
                changed_by=actor,
                before_data=before,
                after_data=shift.to_dict(),
            )
            db.session.commit()
            flash("Đã cập nhật ca", "success")
            return redirect(url_for("edit_shift", shift_id=shift.id))

        return render_template("shift_edit.html", shift=shift)

    @app.route("/shifts/delete/<int:shift_id>", methods=["POST"])
    def delete_shift(shift_id):
        actor = request.form.get("changed_by", "admin").strip() or "admin"

        shift = ShiftTemplate.query.get_or_404(shift_id)
        schedule_count = WorkSchedule.query.filter_by(shift_id=shift.id).count()
        employee_count = Employee.query.filter_by(default_shift_code=shift.code).count()

        if schedule_count > 0 or employee_count > 0:
            flash(
                "Không thể xóa ca này vì đang được sử dụng "
                f"({schedule_count} lịch làm, {employee_count} nhân viên mặc định).",
                "error",
            )
            return redirect(url_for("edit_shift", shift_id=shift.id))

        before = shift.to_dict()
        db.session.delete(shift)
        log_action(
            "shift_templates",
            shift.id,
            "DELETE",
            changed_by=actor,
            before_data=before,
            notes="Xóa mã ca làm",
        )
        db.session.commit()
        flash("Đã xóa mã ca", "success")
        return redirect(url_for("shifts"))

    @app.route("/salaries", methods=["GET", "POST"])
    def salaries():
        month_key = _safe_month_key(request.args.get("month"))
        search_query = (request.args.get("q") or "").strip()
        search_scope = (request.args.get("scope") or "current").strip().lower()
        if search_scope not in {"current", "all"}:
            search_scope = "current"

        if request.method == "POST":
            actor = request.form.get("changed_by", "admin").strip() or "admin"
            action = request.form.get("action", "save_employee_salary").strip()
            target_month = _safe_month_key(request.form.get("month_key") or month_key)

            if action == "save_month_workdays":
                company_work_days = _to_float(request.form.get("company_work_days"), 0)
                notes = request.form.get("notes", "").strip() or None

                if company_work_days <= 0:
                    flash("Công chuẩn tháng phải lớn hơn 0", "error")
                    return redirect(url_for("salaries", month=target_month))

                config = MonthlyWorkdayConfig.query.filter_by(month_key=target_month).first()
                if config:
                    before = config.to_dict()
                    config.company_work_days = company_work_days
                    config.notes = notes
                    log_action(
                        "monthly_workday_configs",
                        config.id,
                        "UPDATE",
                        changed_by=actor,
                        before_data=before,
                        after_data=config.to_dict(),
                    )
                else:
                    config = MonthlyWorkdayConfig(
                        month_key=target_month,
                        company_work_days=company_work_days,
                        notes=notes,
                    )
                    db.session.add(config)
                    db.session.flush()
                    log_action(
                        "monthly_workday_configs",
                        config.id,
                        "INSERT",
                        changed_by=actor,
                        after_data=config.to_dict(),
                    )

                month_salary_rows = MonthlySalary.query.filter_by(month_key=target_month).all()
                for salary_row in month_salary_rows:
                    salary_row.salary_coefficient = company_work_days

                db.session.commit()
                rebuild_month_details(target_month, actor)
                flash("Đã lưu công chuẩn tháng (áp dụng toàn bộ nhân viên)", "success")
                return redirect(url_for("salaries", month=target_month))

            employee_id_raw = request.form.get("employee_id", "").strip()
            if not employee_id_raw.isdigit():
                flash("Cần chọn nhân viên", "error")
                return redirect(url_for("salaries", month=target_month))

            employee_id = int(employee_id_raw)
            base_monthly_wage = _to_float(
                request.form.get("base_monthly_wage") or request.form.get("base_daily_wage"),
                0,
            )
            pay_method = request.form.get("pay_method", "").strip() or None

            if base_monthly_wage < 0:
                flash("Lương tháng không hợp lệ", "error")
                return redirect(url_for("salaries", month=target_month))

            company_work_days, _ = _resolve_company_work_days(target_month)

            row = MonthlySalary.query.filter_by(employee_id=employee_id, month_key=target_month).first()
            if row:
                before = row.to_dict()
                row.base_daily_wage = base_monthly_wage
                row.salary_coefficient = company_work_days
                row.pay_method = pay_method
                log_action(
                    "monthly_salaries",
                    row.id,
                    "UPDATE",
                    changed_by=actor,
                    before_data=before,
                    after_data=row.to_dict(),
                    notes="Lương tháng theo nhân viên",
                )
            else:
                row = MonthlySalary(
                    employee_id=employee_id,
                    month_key=target_month,
                    base_daily_wage=base_monthly_wage,
                    salary_coefficient=company_work_days,
                    pay_method=pay_method,
                )
                db.session.add(row)
                db.session.flush()
                log_action(
                    "monthly_salaries",
                    row.id,
                    "INSERT",
                    changed_by=actor,
                    after_data=row.to_dict(),
                    notes="Lương tháng theo nhân viên",
                )

            db.session.commit()
            rebuild_month_details(target_month, actor)
            flash("Đã lưu lương tháng nhân viên", "success")
            return redirect(url_for("salaries", month=target_month))

        employees = Employee.query.order_by(Employee.employee_code.asc()).all()
        company_work_days_current, workday_config = _resolve_company_work_days(month_key)

        rows_query = MonthlySalary.query.options(joinedload(MonthlySalary.employee))
        if search_scope == "current":
            rows_query = rows_query.filter(MonthlySalary.month_key == month_key)

        if search_query:
            search_like = f"%{search_query}%"
            rows_query = rows_query.join(Employee, MonthlySalary.employee_id == Employee.id).filter(
                or_(
                    MonthlySalary.month_key.ilike(search_like),
                    Employee.employee_code.ilike(search_like),
                    Employee.full_name.ilike(search_like),
                    MonthlySalary.pay_method.ilike(search_like),
                    cast(MonthlySalary.base_daily_wage, String).ilike(search_like),
                    cast(MonthlySalary.salary_coefficient, String).ilike(search_like),
                )
            )

        rows = (
            rows_query
            .order_by(MonthlySalary.month_key.desc(), MonthlySalary.created_at.desc(), MonthlySalary.id.desc())
            .all()
        )

        month_key_set = {row.month_key for row in rows}
        workday_by_month = {}
        if month_key_set:
            workday_rows = MonthlyWorkdayConfig.query.filter(
                MonthlyWorkdayConfig.month_key.in_(month_key_set)
            ).all()
            workday_by_month = {
                row.month_key: _to_float(row.company_work_days, 0)
                for row in workday_rows
            }

        salary_rows = []
        for row in rows:
            company_work_days = workday_by_month.get(row.month_key, 0)
            if company_work_days <= 0:
                legacy_value = _to_float(row.salary_coefficient, 0)
                company_work_days = legacy_value if legacy_value >= 10 else 26.0

            monthly_wage = _to_float(row.base_daily_wage)
            daily_rate = (monthly_wage / company_work_days) if company_work_days > 0 else 0.0
            salary_rows.append(
                {
                    "salary": row,
                    "monthly_wage": round(monthly_wage, 2),
                    "daily_rate": round(daily_rate, 2),
                    "company_work_days": round(company_work_days, 2),
                }
            )

        return render_template(
            "salaries.html",
            month_key=month_key,
            employees=employees,
            salary_rows=salary_rows,
            company_work_days=round(company_work_days_current, 2),
            workday_config=workday_config,
            search_query=search_query,
            search_scope=search_scope,
        )

    @app.route("/salary-overview")
    def salary_overview():
        month_key = _safe_month_key(request.args.get("month"))
        search_query = (request.args.get("q") or "").strip()
        search_scope = (request.args.get("scope") or "current").strip().lower()
        if search_scope not in {"current", "all"}:
            search_scope = "current"

        start_date, end_date = parse_month_key(month_key)
        period_1_end = date(start_date.year, start_date.month, 15)
        period_2_start = date(start_date.year, start_date.month, 16)

        company_work_days_current, _ = _resolve_company_work_days(month_key)
        if company_work_days_current <= 0:
            company_work_days_current = 26.0

        summary_map = {}
        if search_scope == "current":
            employees = Employee.query.filter(Employee.is_active.is_(True)).order_by(Employee.employee_code.asc()).all()
            for row in employees:
                summary_map[(month_key, row.id)] = {
                    "month_key": month_key,
                    "employee": row,
                    "worked_days": 0.0,
                    "paid_leave_days": 0.0,
                    "unpaid_leave_days": 0.0,
                    "overtime_hours": 0.0,
                }

        detail_query = AttendanceDetail.query.options(joinedload(AttendanceDetail.employee))
        if search_scope == "current":
            detail_query = detail_query.filter(AttendanceDetail.month_key == month_key)

        detail_rows = (
            detail_query
            .order_by(
                AttendanceDetail.month_key.desc(),
                AttendanceDetail.employee_id.asc(),
                AttendanceDetail.work_date.asc(),
            )
            .all()
        )

        def _apply_status(summary_obj, status_code):
            normalized = str(status_code or "").upper()
            if normalized == "P":
                summary_obj["paid_leave_days"] += 1.0
            elif normalized in {"S", "C"}:
                summary_obj["paid_leave_days"] += 0.5
                summary_obj["worked_days"] += 0.5
            elif normalized == "N":
                summary_obj["unpaid_leave_days"] += 1.0
            elif normalized == "OFF":
                return
            else:
                summary_obj["worked_days"] += 1.0

        for row in detail_rows:
            if not row.employee:
                continue

            item_month = (row.month_key or month_key).strip()
            summary_key = (item_month, row.employee_id)
            summary = summary_map.get(summary_key)
            if not summary:
                summary = {
                    "month_key": item_month,
                    "employee": row.employee,
                    "worked_days": 0.0,
                    "paid_leave_days": 0.0,
                    "unpaid_leave_days": 0.0,
                    "overtime_hours": 0.0,
                }
                summary_map[summary_key] = summary

            _apply_status(summary, row.status_code)
            summary["overtime_hours"] += _to_float(row.overtime_hours)

        summary_keys = list(summary_map.keys())
        employee_ids = sorted({employee_id for _, employee_id in summary_keys})
        summary_month_keys = sorted({item_month for item_month, _ in summary_keys if item_month})

        salary_by_key = {}
        if employee_ids and summary_month_keys:
            salary_rows = MonthlySalary.query.filter(
                MonthlySalary.employee_id.in_(employee_ids),
                MonthlySalary.month_key.in_(summary_month_keys),
            ).all()
            salary_by_key = {
                (row.month_key, row.employee_id): row
                for row in salary_rows
            }

        advance_by_key = {}
        if employee_ids and summary_month_keys:
            advance_rows = (
                db.session.query(
                    AdvancePayment.month_key,
                    AdvancePayment.employee_id,
                    func.coalesce(func.sum(AdvancePayment.amount), 0),
                )
                .filter(
                    AdvancePayment.month_key.in_(summary_month_keys),
                    AdvancePayment.employee_id.in_(employee_ids),
                )
                .group_by(AdvancePayment.month_key, AdvancePayment.employee_id)
                .all()
            )
            advance_by_key = {
                (item_month, employee_id): _to_float(total_amount)
                for item_month, employee_id, total_amount in advance_rows
            }

        workday_by_month = {month_key: company_work_days_current}
        if search_scope == "all" and summary_month_keys:
            workday_rows = MonthlyWorkdayConfig.query.filter(
                MonthlyWorkdayConfig.month_key.in_(summary_month_keys)
            ).all()
            workday_by_month = {
                row.month_key: _to_float(row.company_work_days, 0)
                for row in workday_rows
            }
            for item_month in summary_month_keys:
                if _to_float(workday_by_month.get(item_month), 0) <= 0:
                    workday_by_month[item_month] = 26.0

        overview_rows = []
        for (item_month, employee_id), summary in summary_map.items():
            salary_row = salary_by_key.get((item_month, employee_id))
            monthly_wage = _to_float(salary_row.base_daily_wage if salary_row else None, 0)

            overtime_hours = round(summary["overtime_hours"], 2)
            overtime_day_units = int(overtime_hours // 8)
            overtime_remainder_hours = round(overtime_hours - (overtime_day_units * 8), 2)

            salary_day_units = (
                summary["worked_days"]
                + summary["paid_leave_days"]
                + (overtime_hours / 8.0)
            )

            company_work_days = _to_float(workday_by_month.get(item_month), 0)
            if company_work_days <= 0:
                legacy_value = _to_float(salary_row.salary_coefficient if salary_row else None, 0)
                company_work_days = legacy_value if legacy_value >= 10 else 26.0

            daily_rate = (monthly_wage / company_work_days) if company_work_days > 0 else 0.0
            salary_amount = round(daily_rate * salary_day_units, 2)
            advance_amount = round(_to_float(advance_by_key.get((item_month, employee_id), 0), 0), 2)

            overview_rows.append(
                {
                    "month_key": item_month,
                    "employee": summary["employee"],
                    "worked_days": round(summary["worked_days"], 2),
                    "paid_leave_days": round(summary["paid_leave_days"], 2),
                    "unpaid_leave_days": round(summary["unpaid_leave_days"], 2),
                    "overtime_day_units": overtime_day_units,
                    "overtime_remainder_hours": overtime_remainder_hours,
                    "advance_amount": advance_amount,
                    "salary_amount": salary_amount,
                }
            )

        if search_scope == "all":
            overview_rows.sort(
                key=lambda item: _employee_code_sort_key(item["employee"].employee_code)
            )
            overview_rows.sort(key=lambda item: item["month_key"], reverse=True)
        else:
            overview_rows.sort(
                key=lambda item: _employee_code_sort_key(item["employee"].employee_code)
            )

        meal_periods = [
            {
                "period": "1",
                "label": "Tiền ăn đợt 1",
                "date_label": f"{start_date.strftime('%d/%m')} - {period_1_end.strftime('%d/%m')}",
            },
            {
                "period": "2",
                "label": "Tiền ăn đợt 2",
                "date_label": f"{period_2_start.strftime('%d/%m')} - {end_date.strftime('%d/%m')}",
            },
        ]

        if search_query:
            search_text = search_query.lower()

            def _match_overview(item):
                values = [
                    item["month_key"],
                    item["employee"].employee_code,
                    item["employee"].full_name,
                    item["worked_days"],
                    item["paid_leave_days"],
                    item["unpaid_leave_days"],
                    item["overtime_day_units"],
                    item["overtime_remainder_hours"],
                    item["advance_amount"],
                    item["salary_amount"],
                ]
                return any(
                    search_text in str(value).lower()
                    for value in values
                    if value is not None
                )

            overview_rows = [item for item in overview_rows if _match_overview(item)]

        return render_template(
            "salary_overview.html",
            month_key=month_key,
            company_work_days=round(company_work_days_current, 2),
            search_query=search_query,
            search_scope=search_scope,
            overview_rows=overview_rows,
            meal_periods=meal_periods,
        )

    @app.route("/salary-overview/meal")
    def salary_overview_meal():
        month_key = _safe_month_key(request.args.get("month"))
        period_raw = (request.args.get("period") or "1").strip()
        period = 2 if period_raw == "2" else 1
        search_query = (request.args.get("q") or "").strip()

        start_date, end_date = parse_month_key(month_key)
        period_1_end = date(start_date.year, start_date.month, 15)
        period_2_start = date(start_date.year, start_date.month, 16)

        if period == 2:
            period_start = period_2_start
            period_end = end_date
            period_label = f"{period_start.strftime('%d/%m')} - {period_end.strftime('%d/%m')}"
            period_title = "Tiền ăn đợt 2"
        else:
            period_start = start_date
            period_end = period_1_end
            period_label = f"{period_start.strftime('%d/%m')} - {period_end.strftime('%d/%m')}"
            period_title = "Tiền ăn đợt 1"

        employees = Employee.query.filter(Employee.is_active.is_(True)).order_by(Employee.employee_code.asc()).all()
        meal_summary_map = {
            row.id: {
                "employee": row,
                "worked_days": 0.0,
                "paid_leave_days": 0.0,
                "unpaid_leave_days": 0.0,
                "meal_amount": 0.0,
            }
            for row in employees
        }

        detail_rows = (
            AttendanceDetail.query.options(joinedload(AttendanceDetail.employee))
            .filter(
                AttendanceDetail.month_key == month_key,
                AttendanceDetail.work_date >= period_start,
                AttendanceDetail.work_date <= period_end,
            )
            .order_by(AttendanceDetail.employee_id.asc(), AttendanceDetail.work_date.asc())
            .all()
        )

        def _apply_status(summary_obj, status_code):
            normalized = str(status_code or "").upper()
            if normalized == "P":
                summary_obj["paid_leave_days"] += 1.0
            elif normalized in {"S", "C"}:
                summary_obj["paid_leave_days"] += 0.5
                summary_obj["worked_days"] += 0.5
            elif normalized == "N":
                summary_obj["unpaid_leave_days"] += 1.0
            elif normalized == "OFF":
                return
            else:
                summary_obj["worked_days"] += 1.0

        for detail_row in detail_rows:
            if not detail_row.employee:
                continue

            meal_summary = meal_summary_map.get(detail_row.employee_id)
            if not meal_summary:
                meal_summary = {
                    "employee": detail_row.employee,
                    "worked_days": 0.0,
                    "paid_leave_days": 0.0,
                    "unpaid_leave_days": 0.0,
                    "meal_amount": 0.0,
                }
                meal_summary_map[detail_row.employee_id] = meal_summary

            _apply_status(meal_summary, detail_row.status_code)
            meal_summary["meal_amount"] += _to_float(detail_row.meal_allowance_daily)

        meal_rows = []
        for employee_id, meal_summary in meal_summary_map.items():
            meal_rows.append(
                {
                    "employee_id": employee_id,
                    "employee": meal_summary["employee"],
                    "worked_days": round(meal_summary["worked_days"], 2),
                    "paid_leave_days": round(meal_summary["paid_leave_days"], 2),
                    "unpaid_leave_days": round(meal_summary["unpaid_leave_days"], 2),
                    "meal_amount": round(meal_summary["meal_amount"], 2),
                }
            )

        meal_rows.sort(
            key=lambda item: _employee_code_sort_key(item["employee"].employee_code)
        )

        if search_query:
            search_text = search_query.lower()

            def _match_meal_row(item):
                values = [
                    item["employee"].employee_code,
                    item["employee"].full_name,
                    item["worked_days"],
                    item["paid_leave_days"],
                    item["unpaid_leave_days"],
                    item["meal_amount"],
                ]
                return any(
                    search_text in str(value).lower()
                    for value in values
                    if value is not None
                )

            meal_rows = [item for item in meal_rows if _match_meal_row(item)]

        return render_template(
            "salary_meal_period.html",
            month_key=month_key,
            period=period,
            period_title=period_title,
            period_label=period_label,
            search_query=search_query,
            meal_rows=meal_rows,
        )

    @app.route("/salary-overview/meal/<int:employee_id>")
    def salary_overview_meal_employee_detail(employee_id):
        month_key = _safe_month_key(request.args.get("month"))
        period_raw = (request.args.get("period") or "1").strip()
        period = 2 if period_raw == "2" else 1
        search_query = (request.args.get("q") or "").strip()

        employee = Employee.query.get_or_404(employee_id)

        start_date, end_date = parse_month_key(month_key)
        period_1_end = date(start_date.year, start_date.month, 15)
        period_2_start = date(start_date.year, start_date.month, 16)

        if period == 2:
            period_start = period_2_start
            period_end = end_date
            period_label = f"{period_start.strftime('%d/%m')} - {period_end.strftime('%d/%m')}"
            period_title = "Tiền ăn đợt 2"
        else:
            period_start = start_date
            period_end = period_1_end
            period_label = f"{period_start.strftime('%d/%m')} - {period_end.strftime('%d/%m')}"
            period_title = "Tiền ăn đợt 1"

        detail_rows = (
            AttendanceDetail.query
            .filter(
                AttendanceDetail.month_key == month_key,
                AttendanceDetail.employee_id == employee.id,
                AttendanceDetail.work_date >= period_start,
                AttendanceDetail.work_date <= period_end,
            )
            .order_by(AttendanceDetail.work_date.asc())
            .all()
        )

        summary = {
            "worked_days": 0.0,
            "paid_leave_days": 0.0,
            "unpaid_leave_days": 0.0,
            "meal_amount": 0.0,
        }

        for row in detail_rows:
            status_code = str(row.status_code or "").upper()
            if status_code == "P":
                summary["paid_leave_days"] += 1.0
            elif status_code in {"S", "C"}:
                summary["paid_leave_days"] += 0.5
                summary["worked_days"] += 0.5
            elif status_code == "N":
                summary["unpaid_leave_days"] += 1.0
            elif status_code == "OFF":
                pass
            else:
                summary["worked_days"] += 1.0

            summary["meal_amount"] += _to_float(row.meal_allowance_daily)

        return render_template(
            "salary_meal_employee_detail.html",
            month_key=month_key,
            period=period,
            period_title=period_title,
            period_label=period_label,
            search_query=search_query,
            employee=employee,
            details=detail_rows,
            summary={
                "worked_days": round(summary["worked_days"], 2),
                "paid_leave_days": round(summary["paid_leave_days"], 2),
                "unpaid_leave_days": round(summary["unpaid_leave_days"], 2),
                "meal_amount": round(summary["meal_amount"], 2),
            },
        )

    @app.route("/salaries/import", methods=["POST"])
    def import_salaries():
        actor = request.form.get("changed_by", "admin").strip() or "admin"
        month_key = _safe_month_key(request.form.get("month_key") or request.args.get("month"))
        replace_existing_month = request.form.get("replace_existing_month") == "on"

        upload = request.files.get("salary_file")
        if not upload or upload.filename == "":
            flash("Cần chọn file hệ lương để import", "error")
            return redirect(url_for("salaries", month=month_key))

        extension = Path(upload.filename).suffix.lower()
        if extension not in {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls", ".csv"}:
            flash("File hệ lương chỉ hỗ trợ CSV/XLSX", "error")
            return redirect(url_for("salaries", month=month_key))

        upload_folder = Path(current_app.config["UPLOAD_FOLDER"])
        upload_folder.mkdir(parents=True, exist_ok=True)

        safe_name = secure_filename(upload.filename)
        temp_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_name}"
        temp_path = upload_folder / temp_name
        upload.save(temp_path)

        try:
            default_work_days, _ = _resolve_company_work_days(month_key)
            result = import_salary_file(
                str(temp_path),
                upload.filename,
                actor=actor,
                target_month=month_key,
                default_company_work_days=default_work_days,
                replace_existing=replace_existing_month,
            )

            rebuild_month_details(month_key, actor)

            replaced_info = ""
            if result["replace_existing"]:
                replaced_info = f" Đã xóa trước {result['deleted_rows']} dòng lương cũ."

            unknown_info = ""
            if result["skipped_unknown"] > 0:
                unknown_preview = ", ".join(result["unknown_codes"][:10])
                unknown_info = (
                    f" Bỏ qua {result['skipped_unknown']} dòng do không tìm thấy Mã NV"
                    f" ({unknown_preview})."
                )

            flash(
                f"Import hệ lương tháng {month_key} xong: tạo mới {result['created']}, "
                f"cập nhật {result['updated']}, công chuẩn {result['company_work_days']}.{replaced_info}{unknown_info}",
                "success",
            )
            return redirect(url_for("salaries", month=month_key))
        except Exception as exc:
            db.session.rollback()
            flash(f"Import hệ lương thất bại: {exc}", "error")
            return redirect(url_for("salaries", month=month_key))
        finally:
            if temp_path.exists():
                os.remove(temp_path)

    @app.route("/advances", methods=["GET", "POST"])
    def advances():
        month_key = _safe_month_key(request.args.get("month") or request.form.get("view_month"))
        search_query = (request.args.get("q") or "").strip()
        search_scope = (request.args.get("scope") or "current").strip().lower()
        if search_scope not in {"current", "all"}:
            search_scope = "current"
        advance_filter = _safe_advance_filter(request.args.get("advance_filter"))

        if request.method == "POST":
            actor = (request.form.get("changed_by") or session.get("username") or "admin").strip() or "admin"
            action = (request.form.get("action") or "create_advance").strip()

            def _build_redirect_params(default_month):
                params = {
                    "month": _safe_month_key(request.form.get("view_month") or default_month),
                }
                view_q = (request.form.get("view_q") or "").strip()
                view_scope = (request.form.get("view_scope") or "current").strip().lower()
                view_filter = _safe_advance_filter(request.form.get("view_advance_filter"))

                if view_q:
                    params["q"] = view_q
                if view_scope == "all":
                    params["scope"] = "all"
                if view_filter != "all":
                    params["advance_filter"] = view_filter
                return params

            if action == "create_advance":
                employee_id_raw = (request.form.get("employee_id") or "").strip()
                if not employee_id_raw.isdigit():
                    flash("Cần chọn nhân viên để thêm ứng tiền", "error")
                    return redirect(url_for("advances", **_build_redirect_params(month_key)))

                employee_id = int(employee_id_raw)
                advance_date_raw = (request.form.get("advance_date") or "").strip()
                payment_method = _safe_payment_method(request.form.get("payment_method"))

                if advance_date_raw:
                    try:
                        advance_date = _parse_date(advance_date_raw)
                    except (TypeError, ValueError):
                        flash("Ngày ứng không hợp lệ", "error")
                        return redirect(url_for("advances", **_build_redirect_params(month_key)))
                else:
                    advance_date = date.today()

                target_month = month_key_for_date(advance_date)
                redirect_params = _build_redirect_params(target_month)
                redirect_params["month"] = target_month

                salary_row = MonthlySalary.query.filter_by(
                    employee_id=employee_id,
                    month_key=target_month,
                ).first()
                if not salary_row:
                    flash(
                        f"Chưa có lương tháng {target_month} cho nhân viên này nên chưa thể ứng tiền",
                        "error",
                    )
                    return redirect(url_for("advances", **redirect_params))

                monthly_wage = _to_float(salary_row.base_daily_wage, 0)
                if monthly_wage <= 0:
                    flash("Lương tháng đang bằng 0, không thể tạo ứng tiền", "error")
                    return redirect(url_for("advances", **redirect_params))

                input_mode = (request.form.get("input_mode") or "amount").strip().lower()
                if input_mode not in {"amount", "days"}:
                    input_mode = "amount"

                advance_days = None
                if input_mode == "days":
                    advance_days = _to_float(request.form.get("advance_days"), 0)
                    if advance_days <= 0:
                        flash("Số ngày công ứng phải lớn hơn 0", "error")
                        return redirect(url_for("advances", **redirect_params))

                    company_work_days, _ = _resolve_company_work_days(target_month)
                    if company_work_days <= 0:
                        company_work_days = 26.0

                    daily_rate = monthly_wage / company_work_days if company_work_days > 0 else 0.0
                    amount = round(daily_rate * advance_days, 2)
                else:
                    amount = _to_float(request.form.get("amount"), 0)

                if amount <= 0:
                    flash("Số tiền ứng phải lớn hơn 0", "error")
                    return redirect(url_for("advances", **redirect_params))

                month_total_existing = (
                    db.session.query(func.coalesce(func.sum(AdvancePayment.amount), 0))
                    .filter(
                        AdvancePayment.employee_id == employee_id,
                        AdvancePayment.month_key == target_month,
                    )
                    .scalar()
                )
                new_total = _to_float(month_total_existing) + amount

                if new_total > monthly_wage + 0.0001:
                    flash(
                        f"Tổng ứng tháng {target_month} ({new_total:,.0f}) vượt lương tháng ({monthly_wage:,.0f}), không thể lưu",
                        "error",
                    )
                    return redirect(url_for("advances", **redirect_params))

                row = AdvancePayment(
                    employee_id=employee_id,
                    advance_date=advance_date,
                    month_key=target_month,
                    amount=amount,
                    input_mode=input_mode,
                    payment_method=payment_method,
                    advance_days=advance_days if input_mode == "days" else None,
                    notes=(request.form.get("notes") or "").strip() or None,
                )
                db.session.add(row)
                db.session.flush()

                log_action(
                    "advance_payments",
                    row.id,
                    "INSERT",
                    changed_by=actor,
                    after_data=row.to_dict(),
                    notes="Tạo giao dịch ứng tiền",
                )

                warning_threshold = monthly_wage * 0.1
                if warning_threshold > 0:
                    if amount >= warning_threshold:
                        flash(
                            f"Cảnh báo: Giao dịch ứng {amount:,.0f} vượt 10% lương tháng ({warning_threshold:,.0f})",
                            "warning",
                        )
                    elif new_total >= warning_threshold:
                        flash(
                            f"Cảnh báo: Tổng ứng tháng hiện tại {new_total:,.0f} vượt 10% lương tháng ({warning_threshold:,.0f})",
                            "warning",
                        )

                db.session.commit()
                flash("Đã thêm giao dịch ứng tiền", "success")
                return redirect(url_for("advances", **redirect_params))

            if action == "update_advance":
                advance_id_raw = (request.form.get("advance_id") or "").strip()
                if not advance_id_raw.isdigit():
                    flash("Không tìm thấy giao dịch ứng tiền để cập nhật", "error")
                    return redirect(url_for("advances", **_build_redirect_params(month_key)))

                row = AdvancePayment.query.get(int(advance_id_raw))
                if not row:
                    flash("Giao dịch ứng tiền không tồn tại", "error")
                    return redirect(url_for("advances", **_build_redirect_params(month_key)))

                advance_date_raw = (request.form.get("advance_date") or "").strip()
                payment_method = _safe_payment_method(request.form.get("payment_method"))
                if advance_date_raw:
                    try:
                        advance_date = _parse_date(advance_date_raw)
                    except (TypeError, ValueError):
                        flash("Ngày ứng không hợp lệ", "error")
                        return redirect(url_for("advances", **_build_redirect_params(row.month_key)))
                else:
                    advance_date = date.today()

                amount = _to_float(request.form.get("amount"), 0)
                if amount <= 0:
                    flash("Số tiền ứng phải lớn hơn 0", "error")
                    return redirect(url_for("advances", **_build_redirect_params(row.month_key)))

                target_month = month_key_for_date(advance_date)
                redirect_params = _build_redirect_params(target_month)
                redirect_params["month"] = target_month

                salary_row = MonthlySalary.query.filter_by(
                    employee_id=row.employee_id,
                    month_key=target_month,
                ).first()
                if not salary_row:
                    flash(
                        f"Chưa có lương tháng {target_month} cho nhân viên này nên chưa thể cập nhật ứng tiền",
                        "error",
                    )
                    return redirect(url_for("advances", **redirect_params))

                monthly_wage = _to_float(salary_row.base_daily_wage, 0)
                if monthly_wage <= 0:
                    flash("Lương tháng đang bằng 0, không thể cập nhật ứng tiền", "error")
                    return redirect(url_for("advances", **redirect_params))

                month_total_existing = (
                    db.session.query(func.coalesce(func.sum(AdvancePayment.amount), 0))
                    .filter(
                        AdvancePayment.employee_id == row.employee_id,
                        AdvancePayment.month_key == target_month,
                        AdvancePayment.id != row.id,
                    )
                    .scalar()
                )
                new_total = _to_float(month_total_existing) + amount

                if new_total > monthly_wage + 0.0001:
                    flash(
                        f"Tổng ứng tháng {target_month} ({new_total:,.0f}) vượt lương tháng ({monthly_wage:,.0f}), không thể lưu",
                        "error",
                    )
                    return redirect(url_for("advances", **redirect_params))

                before = row.to_dict()
                previous_amount = _to_float(row.amount, 0)

                row.advance_date = advance_date
                row.month_key = target_month
                row.amount = amount
                row.payment_method = payment_method
                row.notes = (request.form.get("notes") or "").strip() or None
                if abs(previous_amount - amount) > 0.0001:
                    row.input_mode = "amount"
                    row.advance_days = None

                log_action(
                    "advance_payments",
                    row.id,
                    "UPDATE",
                    changed_by=actor,
                    before_data=before,
                    after_data=row.to_dict(),
                    notes="Cập nhật giao dịch ứng tiền",
                )

                warning_threshold = monthly_wage * 0.1
                if warning_threshold > 0:
                    if amount >= warning_threshold:
                        flash(
                            f"Cảnh báo: Giao dịch ứng {amount:,.0f} vượt 10% lương tháng ({warning_threshold:,.0f})",
                            "warning",
                        )
                    elif new_total >= warning_threshold:
                        flash(
                            f"Cảnh báo: Tổng ứng tháng hiện tại {new_total:,.0f} vượt 10% lương tháng ({warning_threshold:,.0f})",
                            "warning",
                        )

                db.session.commit()
                flash("Đã cập nhật giao dịch ứng tiền", "success")
                return redirect(url_for("advances", **redirect_params))

            if action == "delete_advance":
                advance_id_raw = (request.form.get("advance_id") or "").strip()
                if not advance_id_raw.isdigit():
                    flash("Không tìm thấy giao dịch ứng tiền để xóa", "error")
                    return redirect(url_for("advances", **_build_redirect_params(month_key)))

                row = AdvancePayment.query.get(int(advance_id_raw))
                if not row:
                    flash("Giao dịch ứng tiền không tồn tại", "error")
                    return redirect(url_for("advances", **_build_redirect_params(month_key)))

                target_month = row.month_key
                before = row.to_dict()
                db.session.delete(row)
                log_action(
                    "advance_payments",
                    before.get("id"),
                    "DELETE",
                    changed_by=actor,
                    before_data=before,
                    notes="Xóa giao dịch ứng tiền",
                )
                db.session.commit()
                flash("Đã xóa giao dịch ứng tiền", "success")
                return redirect(url_for("advances", **_build_redirect_params(target_month)))

            flash("Thao tác ứng tiền không hợp lệ", "error")
            return redirect(url_for("advances", **_build_redirect_params(month_key)))

        employees = Employee.query.order_by(Employee.full_name.asc(), Employee.employee_code.asc()).all()

        month_rows = (
            AdvancePayment.query.options(joinedload(AdvancePayment.employee))
            .filter(AdvancePayment.month_key == month_key)
            .order_by(AdvancePayment.advance_date.desc(), AdvancePayment.id.desc())
            .all()
        )

        scope_query = AdvancePayment.query.options(joinedload(AdvancePayment.employee))
        if search_scope == "current":
            scope_query = scope_query.filter(AdvancePayment.month_key == month_key)

        scope_rows = (
            scope_query
            .order_by(
                AdvancePayment.month_key.desc(),
                AdvancePayment.advance_date.desc(),
                AdvancePayment.id.desc(),
            )
            .all()
        )

        month_rows_by_employee = {}
        for row in month_rows:
            month_rows_by_employee.setdefault(row.employee_id, []).append(row)

        scope_rows_by_employee = {}
        for row in scope_rows:
            scope_rows_by_employee.setdefault(row.employee_id, []).append(row)

        search_text = search_query.lower()

        def _advance_row_matches(item):
            payment_method_label = "Theo ngày lương" if item.payment_method == "salary_day" else "Tiền mặt"
            values = [
                item.month_key,
                item.advance_date,
                item.amount,
                item.input_mode,
                item.payment_method,
                payment_method_label,
                item.advance_days,
                item.notes,
            ]
            return any(search_text in str(value).lower() for value in values if value is not None)

        employee_groups = []
        for employee in employees:
            month_items = list(month_rows_by_employee.get(employee.id, []))
            scope_items = list(scope_rows_by_employee.get(employee.id, []))
            visible_items = list(scope_items)

            employee_match = False
            if search_query:
                employee_match = any(
                    search_text in str(value).lower()
                    for value in [employee.employee_code, employee.full_name]
                    if value
                )
                if not employee_match:
                    visible_items = [item for item in scope_items if _advance_row_matches(item)]
                if not employee_match and not visible_items:
                    continue

            has_month_advance = len(month_items) > 0
            if advance_filter == "has" and not has_month_advance:
                continue
            if advance_filter == "none" and has_month_advance:
                continue

            month_total = round(sum(_to_float(item.amount) for item in month_items), 2)
            scope_total = round(sum(_to_float(item.amount) for item in scope_items), 2)
            display_total = round(sum(_to_float(item.amount) for item in visible_items), 2)
            latest_scope_row = scope_items[0] if scope_items else None

            employee_groups.append(
                {
                    "employee": employee,
                    "month_total": month_total,
                    "scope_total": scope_total,
                    "display_total": display_total,
                    "month_count": len(month_items),
                    "scope_count": len(scope_items),
                    "visible_count": len(visible_items),
                    "visible_rows": visible_items,
                    "latest_scope_month": latest_scope_row.month_key if latest_scope_row else None,
                    "latest_scope_date": latest_scope_row.advance_date if latest_scope_row else None,
                }
            )

        employee_groups.sort(
            key=lambda item: _employee_code_sort_key(item["employee"].employee_code)
        )
        employee_groups.sort(
            key=lambda item: item["latest_scope_date"] or date.min,
            reverse=True,
        )
        employee_groups.sort(
            key=lambda item: item["latest_scope_month"] or "",
            reverse=True,
        )

        employees_with_month_advance = sum(1 for item in employee_groups if item["month_count"] > 0)
        month_total_amount = round(sum(item["month_total"] for item in employee_groups), 2)
        scope_total_amount = round(sum(item["scope_total"] for item in employee_groups), 2)
        visible_total_amount = round(sum(item["display_total"] for item in employee_groups), 2)

        return render_template(
            "advances.html",
            month_key=month_key,
            employees=employees,
            employee_groups=employee_groups,
            search_query=search_query,
            search_scope=search_scope,
            advance_filter=advance_filter,
            employees_with_month_advance=employees_with_month_advance,
            month_total_amount=month_total_amount,
            scope_total_amount=scope_total_amount,
            visible_total_amount=visible_total_amount,
            default_advance_date=date.today().isoformat(),
        )

    @app.route("/holidays", methods=["GET", "POST"])
    def holidays():
        month_key = _safe_month_key(
            request.args.get("month") or request.form.get("month_key") or request.form.get("month")
        )
        start_date, end_date = parse_month_key(month_key)

        if request.method == "POST":
            actor = request.form.get("changed_by", "admin").strip() or "admin"
            action = request.form.get("action", "save_single").strip()

            if action == "generate_month":
                created_count = 0
                updated_count = 0
                sunday_total = 0
                sunday_created_count = 0
                vn_created_count = 0
                vn_holidays = _get_vietnam_holiday_map(start_date, end_date)

                current_day = start_date
                while current_day <= end_date:
                    if current_day.weekday() == 6:
                        sunday_total += 1
                        row = Holiday.query.filter_by(holiday_date=current_day).first()
                        if not row:
                            row = Holiday(
                                holiday_date=current_day,
                                name="Chủ nhật",
                                is_paid=True,
                                notes="Tạo tự động theo tháng",
                            )
                            db.session.add(row)
                            db.session.flush()
                            log_action(
                                "holidays",
                                row.id,
                                "INSERT",
                                changed_by=actor,
                                after_data=row.to_dict(),
                                notes="Tạo nhanh Chủ nhật OFF theo tháng",
                            )
                            created_count += 1
                            sunday_created_count += 1
                    current_day += timedelta(days=1)

                for holiday_date, holiday_name in sorted(vn_holidays.items()):
                    row = Holiday.query.filter_by(holiday_date=holiday_date).first()
                    if row:
                        before = row.to_dict()
                        changed = False

                        current_name = (row.name or "").strip()
                        if holiday_name and holiday_name.lower() not in current_name.lower():
                            row.name = f"{current_name} + {holiday_name}" if current_name else holiday_name
                            changed = True

                        if changed:
                            log_action(
                                "holidays",
                                row.id,
                                "UPDATE",
                                changed_by=actor,
                                before_data=before,
                                after_data=row.to_dict(),
                                notes="Bổ sung tên ngày lễ Việt Nam tự động",
                            )
                            updated_count += 1
                    else:
                        row = Holiday(
                            holiday_date=holiday_date,
                            name=holiday_name,
                            is_paid=True,
                            notes="Ngày lễ Việt Nam (tự động theo tháng)",
                        )
                        db.session.add(row)
                        db.session.flush()
                        log_action(
                            "holidays",
                            row.id,
                            "INSERT",
                            changed_by=actor,
                            after_data=row.to_dict(),
                            notes="Tạo tự động ngày lễ Việt Nam",
                        )
                        created_count += 1
                        vn_created_count += 1

                db.session.commit()
                rebuild_month_details(month_key, actor)
                library_note = ""
                if not _get_holiday_library():
                    library_note = " (Đang dùng fallback ngày lễ cố định do chưa cài gói holidays)"

                flash(
                    f"Đã tạo mới {created_count} ngày OFF/lễ (Chủ nhật mới: {sunday_created_count}/{sunday_total}, Lễ VN mới: {vn_created_count}), cập nhật {updated_count} ngày cho tháng {month_key}{library_note}",
                    "success",
                )
                return redirect(url_for("holidays", month=month_key))

            if action == "update_row":
                row_id_raw = (request.form.get("holiday_id") or "").strip()
                if not row_id_raw.isdigit():
                    flash("Không tìm thấy dòng ngày OFF/lễ để cập nhật", "error")
                    return redirect(url_for("holidays", month=month_key))

                row = Holiday.query.get(int(row_id_raw))
                if not row:
                    flash("Dòng ngày OFF/lễ không tồn tại", "error")
                    return redirect(url_for("holidays", month=month_key))

                name = request.form.get("name", "").strip()
                notes = request.form.get("notes", "").strip() or None
                is_paid = request.form.get("is_paid") == "on"

                if not name:
                    flash("Cần nhập tên ngày OFF/lễ", "error")
                    return redirect(url_for("holidays", month=month_key_for_date(row.holiday_date)))

                before = row.to_dict()
                row.name = name
                row.is_paid = is_paid
                row.notes = notes

                log_action(
                    "holidays",
                    row.id,
                    "UPDATE",
                    changed_by=actor,
                    before_data=before,
                    after_data=row.to_dict(),
                )

                db.session.commit()
                rebuild_month_details(month_key_for_date(row.holiday_date), actor)
                flash("Đã cập nhật tick nghỉ/ngày lễ", "success")
                return redirect(url_for("holidays", month=month_key_for_date(row.holiday_date)))

            try:
                holiday_date = _parse_date(request.form.get("holiday_date", "").strip())
            except (TypeError, ValueError):
                flash("Ngày OFF không hợp lệ", "error")
                return redirect(url_for("holidays", month=month_key))

            name = request.form.get("name", "").strip()
            is_paid = request.form.get("is_paid") == "on"
            notes = request.form.get("notes", "").strip() or None

            if not name:
                flash("Cần nhập tên ngày OFF/lễ", "error")
                return redirect(url_for("holidays", month=month_key_for_date(holiday_date)))

            row = Holiday.query.filter_by(holiday_date=holiday_date).first()
            if row:
                before = row.to_dict()
                row.name = name
                row.is_paid = is_paid
                row.notes = notes
                log_action(
                    "holidays",
                    row.id,
                    "UPDATE",
                    changed_by=actor,
                    before_data=before,
                    after_data=row.to_dict(),
                )
            else:
                row = Holiday(holiday_date=holiday_date, name=name, is_paid=is_paid, notes=notes)
                db.session.add(row)
                db.session.flush()
                log_action(
                    "holidays",
                    row.id,
                    "INSERT",
                    changed_by=actor,
                    after_data=row.to_dict(),
                )

            db.session.commit()
            rebuild_month_details(month_key_for_date(holiday_date), actor)
            flash("Đã lưu ngày OFF/lễ", "success")
            return redirect(url_for("holidays", month=month_key_for_date(holiday_date)))

        rows = (
            Holiday.query.filter(
                Holiday.holiday_date >= start_date,
                Holiday.holiday_date <= end_date,
            )
            .order_by(Holiday.holiday_date.asc())
            .all()
        )
        return render_template("holidays.html", holidays=rows, month_key=month_key)

    @app.route("/schedules", methods=["GET", "POST"])
    def schedules():
        month_key = _safe_month_key(request.args.get("month"))

        if request.method == "POST":
            actor = request.form.get("changed_by", "admin")
            employee_id = int(request.form.get("employee_id"))
            work_date = _parse_date(request.form.get("work_date"))
            shift_code = request.form.get("shift_code", "").strip().upper()
            absence_hours = _to_float(request.form.get("absence_hours"), 0)
            overtime_hours_raw = request.form.get("overtime_hours", "").strip()
            overtime_reason = request.form.get("overtime_reason", "").strip() or None
            notes = request.form.get("notes", "").strip() or None

            shift = ShiftTemplate.query.filter_by(code=shift_code).first()
            if not shift:
                flash("Mã ca không hợp lệ", "error")
                return redirect(url_for("schedules", month=month_key_for_date(work_date)))

            row = WorkSchedule.query.filter_by(employee_id=employee_id, work_date=work_date).first()
            target_month = month_key_for_date(work_date)

            if row:
                before = row.to_dict()
                row.shift_id = shift.id
                row.month_key = target_month
                row.absence_hours = absence_hours
                row.notes = notes
                action = "UPDATE"
            else:
                row = WorkSchedule(
                    employee_id=employee_id,
                    work_date=work_date,
                    month_key=target_month,
                    shift_id=shift.id,
                    absence_hours=absence_hours,
                    notes=notes,
                )
                db.session.add(row)
                db.session.flush()
                before = None
                action = "INSERT"

            overtime_hours = (
                _to_float(overtime_hours_raw)
                if overtime_hours_raw
                else float(shift.default_overtime_hours or 0)
            )

            if overtime_hours > 0 or overtime_reason:
                if row.overtime:
                    row.overtime.hours = overtime_hours
                    row.overtime.reason = overtime_reason
                else:
                    overtime = OvertimeEntry(
                        schedule_id=row.id,
                        hours=overtime_hours,
                        reason=overtime_reason,
                    )
                    db.session.add(overtime)

            log_action(
                "work_schedules",
                row.id,
                action,
                changed_by=actor,
                before_data=before,
                after_data=row.to_dict(),
                notes="Lịch làm + tăng ca",
            )

            db.session.commit()
            rebuild_month_details(target_month, actor)
            flash("Đã lưu lịch làm", "success")
            return redirect(url_for("schedules", month=target_month))

        employees = Employee.query.order_by(Employee.employee_code.asc()).all()
        shift_codes = ShiftTemplate.query.order_by(ShiftTemplate.code.asc()).all()

        rows = (
            WorkSchedule.query.options(
                joinedload(WorkSchedule.employee),
                joinedload(WorkSchedule.shift),
                joinedload(WorkSchedule.overtime),
            )
            .filter(WorkSchedule.month_key == month_key)
            .order_by(WorkSchedule.work_date.asc(), WorkSchedule.employee_id.asc())
            .all()
        )

        return render_template(
            "schedules.html",
            month_key=month_key,
            employees=employees,
            schedules=rows,
            shift_codes=shift_codes,
            valid_shift_codes=[row.code for row in shift_codes],
        )

    @app.route("/schedules/import", methods=["POST"])
    def import_schedules():
        actor = request.form.get("changed_by", "admin")
        month_key_input = request.form.get("month_key", "").strip()
        month_key = _safe_month_key(month_key_input) if month_key_input else None
        replace_existing_month = request.form.get("replace_existing_month") == "on"
        open_details_after_import = request.form.get("open_details_after_import") == "on"

        upload = request.files.get("schedule_file")
        if not upload or upload.filename == "":
            flash("Cần chọn file lịch làm .xlsx", "error")
            return redirect(url_for("schedules", month=month_key or current_month_key()))

        extension = Path(upload.filename).suffix.lower()
        if extension not in {".xlsx", ".xlsm", ".xltx", ".xltm", ".xls"}:
            flash("File lịch làm chỉ hỗ trợ định dạng Excel", "error")
            return redirect(url_for("schedules", month=month_key or current_month_key()))

        upload_folder = Path(current_app.config["UPLOAD_FOLDER"])
        upload_folder.mkdir(parents=True, exist_ok=True)

        safe_name = secure_filename(upload.filename)
        temp_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_name}"
        temp_path = upload_folder / temp_name
        upload.save(temp_path)

        try:
            result = import_schedule_file(
                str(temp_path),
                upload.filename,
                actor=actor,
                target_month=month_key,
                replace_existing=replace_existing_month,
            )

            rebuilt = {}
            for item in result["months"]:
                rebuilt[item] = rebuild_month_details(item, actor)

            replaced_info = ""
            if result["replace_existing"] and result["replaced_months"]:
                replaced_info = (
                    f" Đã xóa lịch cũ trước khi import: {result['replaced_months']}."
                )

            flash(
                f"Import lịch xong {result['rows_imported']} dòng ca "
                f"(tạo mới {result['created']}, cập nhật {result['updated']}). "
                f"Đã tái tạo chi tiết: {rebuilt}.{replaced_info}",
                "success",
            )

            redirect_month = month_key or (result["months"][0] if result["months"] else current_month_key())
            if open_details_after_import:
                return redirect(url_for("details", month=redirect_month))
            return redirect(url_for("schedules", month=redirect_month))
        except Exception as exc:
            db.session.rollback()
            flash(f"Import lịch thất bại: {exc}", "error")
            return redirect(url_for("schedules", month=month_key or current_month_key()))
        finally:
            if temp_path.exists():
                os.remove(temp_path)

    @app.route("/imports", methods=["GET", "POST"])
    def imports():
        if request.method == "POST":
            actor = request.form.get("changed_by", "admin")
            month_key_input = request.form.get("month_key", "").strip()
            month_key = _safe_month_key(month_key_input) if month_key_input else None
            replace_existing_month = request.form.get("replace_existing_month") == "on"

            upload = request.files.get("attendance_file")
            if not upload or upload.filename == "":
                flash("Cần chọn file CSV/XLSX", "error")
                return redirect(url_for("imports"))

            upload_folder = Path(current_app.config["UPLOAD_FOLDER"])
            archive_folder = upload_folder / "attendance_imports"
            archive_folder.mkdir(parents=True, exist_ok=True)

            safe_name = secure_filename(upload.filename) or "attendance_upload.csv"
            stored_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}_{safe_name}"
            stored_relpath = str(Path("attendance_imports") / stored_name).replace("\\", "/")
            stored_path = archive_folder / stored_name
            upload.save(stored_path)

            try:
                result = import_attendance_file(
                    str(stored_path),
                    upload.filename,
                    actor,
                    month_key=month_key,
                    replace_existing=replace_existing_month,
                    stored_file_relpath=stored_relpath,
                )
                months = result["months"]
                rebuilt = {}
                for item in months:
                    rebuilt[item] = rebuild_month_details(item, actor)

                replaced_info = ""
                if result["replace_existing"] and result["replaced_months"]:
                    replaced_info = (
                        f" Đã xóa dữ liệu cũ trước khi import: {result['replaced_months']}."
                    )

                flash(
                    f"Import xong {result['rows']} dòng, tạo {result['grouped_days']} bản ghi ngày. "
                    f"Đã tái tạo chi tiết: {rebuilt}.{replaced_info}",
                    "success",
                )
            except Exception as exc:
                db.session.rollback()
                if stored_path.exists():
                    os.remove(stored_path)
                flash(f"Import thất bại: {exc}", "error")

            return redirect(url_for("imports"))

        search_query = (request.args.get("q") or "").strip()
        import_logs_query = AuditLog.query.filter_by(table_name="attendance_import")
        if search_query:
            search_like = f"%{search_query}%"
            import_logs_query = import_logs_query.filter(
                or_(
                    AuditLog.changed_by.ilike(search_like),
                    AuditLog.action.ilike(search_like),
                    AuditLog.record_id.ilike(search_like),
                    cast(AuditLog.after_data, String).ilike(search_like),
                    cast(AuditLog.notes, String).ilike(search_like),
                )
            )

        import_logs = (
            import_logs_query
            .order_by(AuditLog.changed_at.desc())
            .limit(200)
            .all()
        )

        deleted_batch_ids = {
            row[0]
            for row in db.session.query(AuditLog.record_id)
            .filter_by(table_name="attendance_import", action="DELETE_BATCH")
            .all()
            if row[0]
        }

        return render_template(
            "imports.html",
            import_logs=import_logs,
            search_query=search_query,
            deleted_batch_ids=deleted_batch_ids,
        )

    @app.route("/imports/download/<batch_id>", methods=["GET"])
    def download_import_file(batch_id):
        import_log = (
            AuditLog.query.filter_by(
                table_name="attendance_import",
                record_id=batch_id,
                action="IMPORT",
            )
            .order_by(AuditLog.changed_at.desc())
            .first()
        )
        if not import_log:
            flash("Không tìm thấy lô import để tải file", "error")
            return redirect(url_for("imports"))

        payload = import_log.after_data if isinstance(import_log.after_data, dict) else {}
        source_file = str(payload.get("source_file") or f"attendance_{batch_id}.dat")
        stored_file_relpath = payload.get("stored_file")
        stored_file_path = _resolve_upload_relpath(stored_file_relpath)

        if not stored_file_path or not stored_file_path.exists():
            flash("File upload không còn trên máy", "error")
            return redirect(url_for("imports"))

        return send_file(
            str(stored_file_path),
            as_attachment=True,
            download_name=source_file,
        )

    @app.route("/imports/delete/<batch_id>", methods=["POST"])
    def delete_import_batch(batch_id):
        actor = (request.form.get("changed_by") or "admin").strip() or "admin"

        import_log = (
            AuditLog.query.filter_by(
                table_name="attendance_import",
                record_id=batch_id,
                action="IMPORT",
            )
            .order_by(AuditLog.changed_at.desc())
            .first()
        )
        import_payload = import_log.after_data if isinstance(import_log.after_data, dict) else {}
        source_file_name = str(import_payload.get("source_file") or "")
        stored_file_relpath = import_payload.get("stored_file")

        log_rows = AttendanceLog.query.filter_by(import_batch=batch_id).all()
        if not log_rows:
            flash("Không tìm thấy batch import để xóa", "error")
            return redirect(url_for("imports"))

        if not source_file_name and log_rows:
            source_file_name = log_rows[0].source_file or ""

        affected_months = sorted(
            {month_key_for_date(row.event_time.date()) for row in log_rows}
        )

        removed_logs = AttendanceLog.query.filter_by(import_batch=batch_id).delete(
            synchronize_session=False
        )
        removed_daily = AttendanceDaily.query.filter_by(import_batch=batch_id).delete(
            synchronize_session=False
        )

        file_removed = False
        stored_file_path = _resolve_upload_relpath(stored_file_relpath)
        if stored_file_path and stored_file_path.exists():
            try:
                os.remove(stored_file_path)
                file_removed = True
            except OSError:
                file_removed = False

        if import_log:
            import_log.after_data = {
                "source_file": source_file_name,
            }

        log_action(
            "attendance_import",
            batch_id,
            "DELETE_BATCH",
            changed_by=actor,
            after_data={
                "source_file": source_file_name,
                "file_removed": file_removed,
                "removed_logs": int(removed_logs),
                "removed_daily": int(removed_daily),
                "affected_months": affected_months,
            },
            notes="Xóa vĩnh viễn dữ liệu và file upload của một lần import",
        )
        db.session.commit()

        rebuilt = {}
        for item in affected_months:
            rebuilt[item] = rebuild_month_details(item, actor)

        flash(
            f"Đã xóa vĩnh viễn batch {batch_id}. File: {'đã xóa' if file_removed else 'không tìm thấy'}, "
            f"Logs: {removed_logs}, Daily: {removed_daily}, Tái tạo: {rebuilt}",
            "success",
        )
        return redirect(url_for("imports"))

    @app.route("/details")
    def details():
        details_view_data = _collect_details_view_data(request.args, emit_flash=True)
        clear_search_params = dict(details_view_data["query_params"])
        clear_search_params.pop("q", None)
        clear_search_params.pop("scope", None)

        employees = Employee.query.order_by(Employee.employee_code.asc()).all()
        return render_template(
            "details.html",
            month_key=details_view_data["month_key"],
            details=details_view_data["rows"],
            search_query=details_view_data["search_query"],
            search_scope=details_view_data["search_scope"],
            employees=employees,
            selected_employee_id=details_view_data["selected_employee_id"],
            start_date_value=(
                details_view_data["parsed_start_date"].isoformat()
                if details_view_data["parsed_start_date"]
                else ""
            ),
            end_date_value=(
                details_view_data["parsed_end_date"].isoformat()
                if details_view_data["parsed_end_date"]
                else ""
            ),
            is_range_mode=details_view_data["is_range_mode"],
            period_label=details_view_data["period_label"],
            clear_search_url=url_for("details", **clear_search_params),
            export_excel_url=url_for("export_details_excel", **details_view_data["query_params"]),
        )

    @app.route("/details/export.xlsx")
    def export_details_excel():
        details_view_data = _collect_details_view_data(request.args, emit_flash=False)
        rows = details_view_data["rows"]

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Bảng chi tiết"

        headers = [
            "STT",
            "Mã NV",
            "Họ tên",
            "Ngày",
            "Tên ca",
            "Giờ vào",
            "Giờ ra",
            "Giờ thực",
            "Chênh lệch",
            "Tăng ca",
            "Tổng giờ",
            "Mã trạng thái",
            "Số giờ",
            "Lương theo ngày",
            "Ghi chú",
            "Tiền ăn theo ngày",
        ]
        sheet.append(headers)

        fill_by_tag = {
            key: PatternFill(fill_type="solid", fgColor=value)
            for key, value in DETAILS_HIGHLIGHT_TO_EXCEL_FILL.items()
        }

        for index, row in enumerate(rows, start=1):
            sheet.append(
                [
                    index,
                    row.employee.employee_code,
                    row.employee.full_name,
                    row.work_date.isoformat(),
                    f"{row.shift_code} - {row.shift_name}",
                    row.check_in.strftime("%Y-%m-%d %H:%M") if row.check_in else "",
                    row.check_out.strftime("%Y-%m-%d %H:%M") if row.check_out else "",
                    float(row.actual_work_hours),
                    float(row.deviation_hours),
                    float(row.overtime_hours),
                    float(row.total_span_hours),
                    row.status_code,
                    float(row.paid_hours),
                    float(row.daily_wage),
                    row.notes or "",
                    float(row.meal_allowance_daily),
                ]
            )

            highlight_tag = getattr(row, "highlight_tag", "") or _get_details_highlight_tag(row)
            row_fill = fill_by_tag.get(highlight_tag)
            if row_fill:
                for cell in sheet[sheet.max_row]:
                    cell.fill = row_fill

        sheet.freeze_panes = "A2"

        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)

        if details_view_data["is_range_mode"]:
            start_date_label = details_view_data["parsed_start_date"].strftime("%Y%m%d")
            end_date_label = details_view_data["parsed_end_date"].strftime("%Y%m%d")
            filename = f"bang_chi_tiet_{start_date_label}_{end_date_label}.xlsx"
        else:
            filename = f"bang_chi_tiet_{details_view_data['month_key'].replace('-', '')}.xlsx"

        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    @app.route("/audit")
    def audit_logs():
        rows = AuditLog.query.order_by(AuditLog.changed_at.desc()).limit(300).all()
        return render_template("audit.html", audit_logs=rows)

    @app.route("/audit/export")
    def export_audit_csv():
        rows = AuditLog.query.order_by(AuditLog.changed_at.desc()).limit(5000).all()

        stream = io.StringIO()
        writer = csv.writer(stream)
        writer.writerow(
            [
                "ID",
                "Thời gian",
                "Người sửa",
                "Bảng",
                "Hành động",
                "Bản ghi",
                "Trước",
                "Sau",
                "Ghi chú",
            ]
        )

        for row in rows:
            writer.writerow(
                [
                    row.id,
                    row.changed_at,
                    row.changed_by,
                    row.table_name,
                    row.action,
                    row.record_id,
                    row.before_data,
                    row.after_data,
                    row.notes,
                ]
            )

        output = stream.getvalue()
        stream.close()

        return Response(
            output,
            mimetype="text/csv",
            headers={
                "Content-Disposition": "attachment; filename=audit_logs.csv",
            },
        )

    @app.route("/backup/run", methods=["POST"])
    def run_backup_now():
        actor = request.form.get("changed_by", "admin")
        try:
            backup_file, removed = run_pg_dump(
                current_app.config["SQLALCHEMY_DATABASE_URI"],
                current_app.config["BACKUP_TARGET_DIR"],
                current_app.config["BACKUP_RETENTION_DAYS"],
            )
            log_action(
                "system_backup",
                backup_file,
                "BACKUP",
                changed_by=actor,
                after_data={"backup_file": backup_file, "removed_files": removed},
                notes="Backup thủ công",
            )
            db.session.commit()
            flash(f"Backup thành công: {backup_file}", "success")
        except Exception as exc:
            db.session.rollback()
            flash(f"Backup thất bại: {exc}", "error")

        return redirect(url_for("audit_logs"))
