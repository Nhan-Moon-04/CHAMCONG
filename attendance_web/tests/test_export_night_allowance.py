import sys
sys.path.insert(0, 'attendance_web')
from app.services.salary_meal_export import build_salary_meal_export_excel
from openpyxl import load_workbook

E = type('E', (), {})
# employee A no night shifts
emp_a = E(); emp_a.employee_code = 'A1'; emp_a.full_name = 'Alice'
# employee B has night shifts
emp_b = E(); emp_b.employee_code = 'B1'; emp_b.full_name = 'Betty'

sample = {
    'period': 1,
    'month_key': '2026-04',
    'period_start': None,
    'period_end': None,
    'meal_rows': [
        {'employee': emp_a, 'meal_count': 5, 'meal_allowance': 40000, 'night_shift_count': 0, 'worked_days':5, 'paid_leave_days':0, 'unpaid_leave_days':0, 'meal_amount':200000},
        {'employee': emp_b, 'meal_count': 3, 'meal_allowance': 35000, 'night_shift_count': 2, 'worked_days':3, 'paid_leave_days':0, 'unpaid_leave_days':0, 'meal_amount':130000},
    ]
}

out, fn = build_salary_meal_export_excel(sample)
wb = load_workbook(out, data_only=False)
ws = wb.active
# header row is 5, data rows start at 6
print('headers:', [c.value for c in ws[5]])
print('row1 night allowance (col8):', ws.cell(6,8).value)
print('row1 night count (col7):', ws.cell(6,7).value)
print('row2 night allowance (col8):', ws.cell(7,8).value)
print('row2 night count (col7):', ws.cell(7,7).value)
print('last header (col15):', ws.cell(5,15).value)
